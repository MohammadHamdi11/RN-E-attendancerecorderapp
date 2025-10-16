import sys
import re
import pandas as pd
import traceback
import os
import requests
import json
import random
import shutil
import threading
from collections import defaultdict
from datetime import datetime, timedelta, date, time
import pytz
from typing import List, Dict
import io
from PIL import Image
import math
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import glob




#==========================================================Session Analyzer==========================================================#

class SessionAnalyzer:
    """Independent session analyzer class that can be used by the processor"""
    
    def __init__(self, parent_widget=None):
        self.parent_widget = parent_widget  # Store parent widget for message boxes
        self.schedule_data = []
        self.log_data = []
        self.reference_data = []
        self.missing_sessions = []
        self.recorded_sessions = []
        
    def get_egypt_time(self):
        """Get current time in Egypt timezone using timezone-aware datetime"""
        # Use timezone-aware datetime instead of deprecated utcnow()
        from datetime import timezone
        utc_now = datetime.now(timezone.utc)
        
        # Egypt is UTC+2 (or UTC+3 during daylight saving)
        # Using pytz is more reliable for handling DST automatically
        egypt_tz = pytz.timezone('Africa/Cairo')
        egypt_time = utc_now.astimezone(egypt_tz)
        
        return egypt_time
        
    def normalize_whitespace(self, text):
        """
        Normalize whitespace in text by:
        1. Removing leading whitespace
        2. Removing trailing whitespace
        3. Collapsing multiple consecutive spaces into one
        """
        if not isinstance(text, str):
            return text
        
        # Strip leading and trailing whitespace
        text = text.strip()
        
        # Replace multiple consecutive spaces with single space
        text = ' '.join(text.split())
        
        return text
    
    def analyze_sessions_with_merged_logs(self, ref_file_path, ref_sheet, merged_log_data, 
                           schedule_file_path, schedule_sheet, year, batch, module):
        """Analyze sessions using already merged log data and identify missing/recorded ones with group-specific attendance calculation"""
        try:
            # Load reference data
            if not self.load_reference_data(ref_file_path, ref_sheet):
                return False, "Failed to load reference data", []
            
            # Load schedule data
            schedule_wb = openpyxl.load_workbook(schedule_file_path, read_only=True)
            schedule_ws = schedule_wb[schedule_sheet]
            schedule_data = list(schedule_ws.values)
            
            # Use merged log data directly instead of loading from file
            log_data = merged_log_data
            
            # Parse schedule data (skip header row)
            self.schedule_data = []
            for row in schedule_data[1:]:
                if len(row) >= 6:  # Ensure we have required fields
                    year_val, group, subject, session_num, date_val, start_time = row[:6]
                    if year_val and group and subject and session_num and date_val:
                        # NORMALIZE AND CONVERT TO UPPERCASE FOR CASE-INSENSITIVE MATCHING
                        year_val = self.normalize_whitespace(str(year_val)) if year_val else ""
                        group = self.normalize_whitespace(str(group).upper()) if group else ""
                        subject = self.normalize_whitespace(str(subject).upper()) if subject else ""
                        session_num = self.normalize_whitespace(str(session_num)) if session_num else ""
                        
                        self.schedule_data.append({
                            'year': str(year_val),
                            'group': group,
                            'subject': subject,
                            'session': str(session_num),
                            'date': self.normalize_date(date_val),
                            'start_time': self.normalize_time(start_time),
                            'raw_date': date_val,
                            'raw_time': start_time
                        })
            
            # Parse log data (skip header row if it exists)
            self.log_data = []
            start_idx = 1 if log_data and len(log_data) > 0 and log_data[0] and any(isinstance(cell, str) and 'Student' in str(cell) for cell in log_data[0][:2]) else 0
            
            for row in log_data[start_idx:]:
                if len(row) >= 6:  # Ensure we have required fields
                    student_id, location, log_date, log_time, type_field, user = row[:6]
                    if location and log_date:  # Use location instead of subject for log data
                        # CONVERT TO UPPERCASE FOR CASE-INSENSITIVE MATCHING
                        location = str(location).upper() if location else ""
                        
                        self.log_data.append({
                            'student_id': str(student_id) if student_id else '',
                            'subject': location,  # Map location to subject for matching
                            'log_date': self.normalize_date(log_date),
                            'log_time': self.normalize_time(log_time),
                            'type': str(type_field) if type_field else '',
                            'user': str(user) if user else '',
                            'raw_date': log_date,
                            'raw_time': log_time
                        })
            
            # Find missing and recorded sessions with GROUP-SPECIFIC attendance
            self.missing_sessions = []
            self.recorded_sessions = []
            
            for scheduled_session in self.schedule_data:
                # Look for matching log entries (same subject and date)
                matching_logs = []
                unique_users = set()        # Track unique users who recorded this session
                
                for log_entry in self.log_data:
                    if (log_entry['subject'] == scheduled_session['subject'] and
                        log_entry['log_date'] == scheduled_session['date']):
                        matching_logs.append(log_entry)
                        
                        # Add to unique users set
                        if log_entry['user']:
                            unique_users.add(log_entry['user'])
                
                if matching_logs:
                    # Session was recorded - calculate GROUP-SPECIFIC attendance
                    group_attendance = self.get_group_specific_attendance(scheduled_session, matching_logs)
                    
                    # Only mark as "Recorded" if at least one valid student attended
                    if group_attendance['unique_student_count'] > 0:
                        self.recorded_sessions.append({
                            **scheduled_session,
                            'recorded_by': list(unique_users),
                            'unique_student_count': group_attendance['unique_student_count'],
                            'expected_student_count': group_attendance['expected_student_count'],
                            'attendance_rate': group_attendance['attendance_rate'],
                            'total_entry_count': len(matching_logs),
                            'group_specific_entry_count': len(group_attendance['group_specific_logs']),
                            'attending_student_ids': group_attendance['attending_student_ids']
                        })
                    else:
                        # Logs exist but no valid students - treat as missing
                        expected_students = self.get_expected_students_for_group(
                            scheduled_session['year'], 
                            scheduled_session['group']
                        )
                        expected_count = len(expected_students)
                        
                        session_copy = scheduled_session.copy()
                        session_copy['expected_student_count'] = expected_count
                        session_copy['warning'] = f"Logs found ({len(matching_logs)}) but no valid students from this group"
                        self.missing_sessions.append(session_copy)
                else:
                    # ← ADD THIS ELSE BLOCK
                    # Session is completely missing (no logs at all)
                    expected_students = self.get_expected_students_for_group(
                        scheduled_session['year'], 
                        scheduled_session['group']
                    )
                    expected_count = len(expected_students)
                    
                    session_copy = scheduled_session.copy()
                    session_copy['expected_student_count'] = expected_count
                    self.missing_sessions.append(session_copy)
            
            # Export results and return generated file paths (xlsx and json)
            generated_files = self.export_results(year, batch, module)
            
            return True, "Analysis completed successfully.", generated_files
            
        except Exception as e:
            return False, f"Error during analysis: {str(e)}", []
    
    def analyze_sessions(self, ref_file_path, ref_sheet, log_file_path, log_sheet, 
                       schedule_file_path, schedule_sheet, year, batch, module):
        """Analyze sessions and identify missing/recorded ones with group-specific attendance calculation"""
        try:
            # Load reference data
            if not self.load_reference_data(ref_file_path, ref_sheet):
                return False, "Failed to load reference data", []
            
            # Load schedule data
            schedule_wb = openpyxl.load_workbook(schedule_file_path, read_only=True)
            schedule_ws = schedule_wb[schedule_sheet]
            schedule_data = list(schedule_ws.values)
            
            # Load log data
            log_wb = openpyxl.load_workbook(log_file_path, read_only=True)
            log_ws = log_wb[log_sheet]
            log_data = list(log_ws.values)
            
            # Parse schedule data (skip header row)
            self.schedule_data = []
            for row in schedule_data[1:]:
                if len(row) >= 6:  # Ensure we have required fields
                    year_val, group, subject, session_num, date_val, start_time = row[:6]
                    if year_val and group and subject and session_num and date_val:
                        # CONVERT TO UPPERCASE FOR CASE-INSENSITIVE MATCHING
                        group = str(group).upper() if group else ""
                        subject = str(subject).upper() if subject else ""
                        
                        self.schedule_data.append({
                            'year': str(year_val),
                            'group': group,
                            'subject': subject,
                            'session': str(session_num),
                            'date': self.normalize_date(date_val),
                            'start_time': self.normalize_time(start_time),
                            'raw_date': date_val,
                            'raw_time': start_time
                        })
            
            # Parse log data (skip header row)
            self.log_data = []
            for row in log_data[1:]:
                if len(row) >= 6:  # Ensure we have required fields
                    student_id, subject, log_date, log_time, type_field, user = row[:6]
                    if subject and log_date:
                        # NORMALIZE AND CONVERT TO UPPERCASE FOR CASE-INSENSITIVE MATCHING
                        student_id = self.normalize_whitespace(str(student_id)) if student_id else ''
                        subject = self.normalize_whitespace(str(subject).upper()) if subject else ""
                        type_field = self.normalize_whitespace(str(type_field)) if type_field else ''
                        user = self.normalize_whitespace(str(user)) if user else ''
                        
                        self.log_data.append({
                            'student_id': str(student_id) if student_id else '',
                            'subject': subject,
                            'log_date': self.normalize_date(log_date),
                            'log_time': self.normalize_time(log_time),
                            'type': str(type_field) if type_field else '',
                            'user': str(user) if user else '',
                            'raw_date': log_date,
                            'raw_time': log_time
                        })
            
            # Find missing and recorded sessions with GROUP-SPECIFIC attendance
            self.missing_sessions = []
            self.recorded_sessions = []
            
            for scheduled_session in self.schedule_data:
                # Look for matching log entries (same subject and date)
                matching_logs = []
                unique_users = set()        # Track unique users who recorded this session
                
                for log_entry in self.log_data:
                    if (log_entry['subject'] == scheduled_session['subject'] and
                        log_entry['log_date'] == scheduled_session['date']):
                        matching_logs.append(log_entry)
                        
                        # Add to unique users set
                        if log_entry['user']:
                            unique_users.add(log_entry['user'])
                
                if matching_logs:
                    # Session was recorded - calculate GROUP-SPECIFIC attendance
                    group_attendance = self.get_group_specific_attendance(scheduled_session, matching_logs)
                    
                    # Only mark as "Recorded" if at least one valid student attended
                    if group_attendance['unique_student_count'] > 0:
                        self.recorded_sessions.append({
                            **scheduled_session,
                            'recorded_by': list(unique_users),
                            'unique_student_count': group_attendance['unique_student_count'],
                            'expected_student_count': group_attendance['expected_student_count'],
                            'attendance_rate': group_attendance['attendance_rate'],
                            'total_entry_count': len(matching_logs),
                            'group_specific_entry_count': len(group_attendance['group_specific_logs']),
                            'attending_student_ids': group_attendance['attending_student_ids']
                        })
                    else:
                        # Logs exist but no valid students - treat as missing
                        expected_students = self.get_expected_students_for_group(
                            scheduled_session['year'], 
                            scheduled_session['group']
                        )
                        expected_count = len(expected_students)
                        
                        session_copy = scheduled_session.copy()
                        session_copy['expected_student_count'] = expected_count
                        session_copy['warning'] = f"Logs found ({len(matching_logs)}) but no valid students from this group"
                        self.missing_sessions.append(session_copy)
                else:
                    # ← ADD THIS ELSE BLOCK
                    # Session is completely missing (no logs at all)
                    expected_students = self.get_expected_students_for_group(
                        scheduled_session['year'], 
                        scheduled_session['group']
                    )
                    expected_count = len(expected_students)
                    
                    session_copy = scheduled_session.copy()
                    session_copy['expected_student_count'] = expected_count
                    self.missing_sessions.append(session_copy)
            
            # Export results and return generated file paths (xlsx and json)
            generated_files = self.export_results(year, batch, module)
            
            return True, "Analysis completed successfully.", generated_files
            
        except Exception as e:
            return False, f"Error during analysis: {str(e)}", []
    
    def load_reference_data(self, ref_file_path, ref_sheet):
        """Load reference data containing student enrollment information"""
        try:
            # Load reference data
            reference_wb = openpyxl.load_workbook(ref_file_path, read_only=True)
            reference_ws = reference_wb[ref_sheet]
            reference_data = list(reference_ws.values)
            
            # Parse reference data (skip header row)
            self.reference_data = []
            for row in reference_data[1:]:
                if len(row) >= 4:  # Ensure we have required fields: Student ID, Name, Year, Group
                    student_id, name, year, group = row[:4]
                    if student_id and group:  # At minimum we need Student ID and Group
                        # NORMALIZE AND CONVERT TO UPPERCASE FOR CASE-INSENSITIVE MATCHING
                        student_id = self.normalize_whitespace(str(student_id))
                        name = self.normalize_whitespace(str(name)) if name else ""
                        year = self.normalize_whitespace(str(year)) if year else ""
                        group = self.normalize_whitespace(str(group).upper()) if group else ""
                        
                        self.reference_data.append({
                            'student_id': str(student_id),
                            'name': str(name) if name else "",
                            'year': str(year) if year else "",
                            'group': group
                        })
            
            return True
        except Exception as e:
            return False
    
    def get_expected_students_for_group(self, year, group):
        """Get the list of expected students for a specific year and group from reference data"""
        # Normalize inputs for matching
        year = self.normalize_whitespace(str(year))
        group = self.normalize_whitespace(str(group).upper())
        
        expected_students = []
        for student in self.reference_data:
            if (student['year'] == year and student['group'] == group):
                expected_students.append(student)
        return expected_students
    
    def get_group_specific_attendance(self, session, matching_logs):
        """Calculate attendance specifically for the session's group"""
        # Get expected students for this group
        expected_students = self.get_expected_students_for_group(session['year'], session['group'])
        expected_student_ids = {student['student_id'] for student in expected_students}
        
        # Filter logs to only include students from the expected group
        group_specific_logs = []
        for log in matching_logs:
            if log['student_id'] in expected_student_ids:
                group_specific_logs.append(log)
        
        # Get unique student IDs from group-specific logs
        attending_student_ids = set()
        for log in group_specific_logs:
            if log['student_id']:
                attending_student_ids.add(log['student_id'])
        
        unique_student_count = len(attending_student_ids)
        expected_student_count = len(expected_student_ids)
        attendance_rate = (unique_student_count / expected_student_count * 100) if expected_student_count > 0 else 0
        
        return {
            'unique_student_count': unique_student_count,
            'expected_student_count': expected_student_count,
            'attendance_rate': attendance_rate,
            'group_specific_logs': group_specific_logs,
            'attending_student_ids': list(attending_student_ids),
            'total_logs_for_session': len(matching_logs)
        }
    
    def normalize_date(self, date_value):
        """Normalize date to datetime object"""
        if isinstance(date_value, datetime):
            return date_value.date()
        elif isinstance(date_value, date):
            return date_value
        elif isinstance(date_value, str):
            try:
                # Try parsing common date formats
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                    try:
                        return datetime.strptime(date_value, fmt).date()
                    except ValueError:
                        continue
            except:
                pass
        return None
    
    def normalize_time(self, time_value):
        """Normalize time to time object"""
        if isinstance(time_value, time):
            return time_value
        elif isinstance(time_value, datetime):
            return time_value.time()
        elif isinstance(time_value, str):
            try:
                # Try parsing common time formats
                for fmt in ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p']:
                    try:
                        return datetime.strptime(time_value, fmt).time()
                    except ValueError:
                        continue
            except:
                pass
        return None
    
    def export_results(self, year, batch, module):
        """Export analysis results to both Excel and JSON files with consistent naming"""
        if not hasattr(self, 'missing_sessions') or not hasattr(self, 'recorded_sessions'):
            return []

        try:
            # Use the same directory structure as the main processor
            base_dir = os.getcwd()
            output_dir = os.path.join(base_dir, "attendance_reports")
            os.makedirs(output_dir, exist_ok=True)

            # Generate filename using the same pattern as attendance reports
            # Format: Y{year}_B{batch}_{module}_session_analysis
            now = datetime.now()
            
            # Build filename components
            filename_parts = []
            if year and year != "Unknown":
                filename_parts.append(f"Y{year}")
            if batch and batch != "Unknown":
                # Handle batch formatting (e.g., "2425" -> "B2425")
                if not batch.startswith('B'):
                    filename_parts.append(f"B{batch}")
                else:
                    filename_parts.append(batch)
            if module and module != "Unknown_Module":
                filename_parts.append(module)
            
            # Add session analysis suffix
            filename_parts.append("session_analysis")
            
            if filename_parts:
                base_filename = '_'.join(filename_parts)
            else:
                # Fallback naming
                date_str = now.strftime('%d%m%Y')
                time_str = now.strftime('%H%M%S')
                base_filename = f"Session_Analysis_Results_D{date_str}T{time_str}"
            
            # File paths for both Excel and JSON
            excel_filepath = os.path.join(output_dir, base_filename + ".xlsx")
            json_filepath = os.path.join(output_dir, base_filename + ".json")

            # Get current date and time for pending status determination
            current_datetime = self.get_egypt_time()
            current_date = current_datetime.date()
            current_time = current_datetime.time()
            
            # Combine all sessions with their status (this data will be in both Excel and JSON)
            all_sessions = []
            
            # Add recorded sessions
            for session in self.recorded_sessions:
                session_copy = session.copy()
                session_copy['status'] = 'Recorded'
                session_copy['unique_student_count'] = session.get('unique_student_count', 0)
                session_copy['expected_student_count'] = session.get('expected_student_count', 0)
                session_copy['attendance_rate'] = session.get('attendance_rate', 0)
                session_copy['recorded_by'] = ', '.join(session.get('recorded_by', ['Unknown']))
                session_copy['total_entry_count'] = session.get('total_entry_count', 0)
                all_sessions.append(session_copy)
            
            # Add missing sessions and determine if they are "Not Recorded" or "Pending"
            for session in self.missing_sessions:
                session_copy = session.copy()
                session_copy['unique_student_count'] = 0
                session_copy['expected_student_count'] = session.get('expected_student_count', 0)
                session_copy['attendance_rate'] = 0
                session_copy['recorded_by'] = ''
                session_copy['total_entry_count'] = 0
                
                # Determine status based on current date/time
                session_date = session['date']
                session_time = session.get('start_time')
                
                # Convert session_date to date object if it's not already
                if isinstance(session_date, str):
                    try:
                        parsed_date = datetime.strptime(session_date, '%Y-%m-%d')
                        session_date = parsed_date.date()
                    except:
                        try:
                            parsed_date = datetime.strptime(session_date, '%d/%m/%Y')
                            session_date = parsed_date.date()
                        except:
                            session_date = current_date  # Fallback
                
                # Determine if session is pending or not recorded
                if session_date > current_date:
                    session_copy['status'] = 'Pending'
                elif session_date == current_date:
                    if session_time and hasattr(session_time, 'hour'):
                        if session_time > current_time:
                            session_copy['status'] = 'Pending'
                        else:
                            session_copy['status'] = 'Not Recorded'
                    else:
                        session_copy['status'] = 'Not Recorded'
                else:
                    session_copy['status'] = 'Not Recorded'
                
                all_sessions.append(session_copy)
            
            # Sort sessions by year, group, subject, session, date, time
            def sort_key(session):
                try:
                    session_num = int(session.get('session', 0))
                except (ValueError, TypeError):
                    session_num = 0
                
                return (
                    str(session.get('year', '')),
                    str(session.get('group', '')),
                    str(session.get('subject', '')),
                    session_num,
                    session.get('date', datetime.min.date()),
                    session.get('start_time', datetime.min.time()),
                    str(session.get('recorded_by', ''))
                )
            
            all_sessions.sort(key=sort_key)
            
            # === JSON EXPORT (Main session data only) ===
            # Prepare JSON data (convert non-serializable objects to strings)
            json_sessions = []
            for session in all_sessions:
                json_session = {}
                for key, value in session.items():
                    if hasattr(value, 'strftime'):  # datetime, date, time objects
                        if key in ['raw_date', 'date']:
                            json_session[key] = value.strftime('%d/%m/%Y') if hasattr(value, 'strftime') else str(value)
                        elif key in ['raw_time', 'start_time']:
                            json_session[key] = value.strftime('%H:%M:%S') if hasattr(value, 'strftime') else str(value)
                        else:
                            json_session[key] = str(value)
                    elif isinstance(value, list):
                        json_session[key] = value  # Lists are JSON serializable
                    else:
                        json_session[key] = value
                json_sessions.append(json_session)
            
            # Create JSON data structure (sessions only, no statistics tables)
            json_data = {
                'metadata': {
                    'export_timestamp': now.isoformat(),
                    'year': year,
                    'batch': batch,
                    'module': module,
                    'export_note': 'Contains main session data only, without statistics tables'
                },
                'sessions': json_sessions
            }
            
            # Save JSON file
            with open(json_filepath, 'w', encoding='utf-8') as json_file:
                json.dump(json_data, json_file, indent=2, ensure_ascii=False)
            
            # === EXCEL EXPORT (With all tables and statistics) ===
            # Create Excel file
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create main results sheet
            main_ws = wb.create_sheet("Session Analysis Results")
            
            # Headers with group-specific columns
            headers = [
                'Year', 'Group', 'Subject', 'Session', 'Date', 'Time', 
                'Recorded By', 'Students', 'Status'
            ]
            main_ws.append(headers)
            
            # Style headers
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, cell in enumerate(main_ws[1], 1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Define colors for status
            recorded_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            not_recorded_fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            pending_fill = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
            
            # Add data rows
            for row_num, session in enumerate(all_sessions, 2):
                # Format date and time for display
                date_str = str(session.get('raw_date', session.get('date', '')))
                time_str = str(session.get('raw_time', session.get('start_time', '')))
                
                if hasattr(session.get('raw_date'), 'strftime'):
                    date_str = session['raw_date'].strftime('%d/%m/%Y')
                elif hasattr(session.get('date'), 'strftime'):
                    date_str = session['date'].strftime('%d/%m/%Y')
                
                if hasattr(session.get('raw_time'), 'strftime'):
                    time_str = session['raw_time'].strftime('%H:%M:%S')
                elif hasattr(session.get('start_time'), 'strftime'):
                    time_str = session['start_time'].strftime('%H:%M:%S')
                
                # Prepare row data with attending/total format
                row_data = [
                    str(session.get('year', '')),
                    str(session.get('group', '')),
                    str(session.get('subject', '')),
                    str(session.get('session', '')),
                    date_str,
                    time_str,
                    str(session.get('recorded_by', '')),
                    f"{session.get('unique_student_count', 0)}/{session.get('expected_student_count', 0)}",
                    session.get('status', '')
                ]
                
                main_ws.append(row_data)
                
                # Apply color coding based on status with black font
                status = session.get('status', '')
                
                for col_num in range(1, len(headers) + 1):
                    cell = main_ws.cell(row=row_num, column=col_num)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Set font to black for better readability
                    cell.font = Font(color='000000')
                    
                    # Apply status-based coloring to entire row
                    if status == 'Recorded':
                        cell.fill = recorded_fill
                    elif status == 'Not Recorded':
                        cell.fill = not_recorded_fill
                    elif status == 'Pending':
                        cell.fill = pending_fill
            
            # Auto-size columns for main data
            for column in main_ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                main_ws.column_dimensions[column_letter].width = adjusted_width
            
            # Calculate overall class statistics
            total_sessions = len(all_sessions)
            recorded_sessions = len([s for s in all_sessions if s.get('status') == 'Recorded'])
            missing_sessions = len([s for s in all_sessions if s.get('status') == 'Not Recorded'])
            pending_sessions = len([s for s in all_sessions if s.get('status') == 'Pending'])
            coverage_percentage = (recorded_sessions / total_sessions * 100) if total_sessions > 0 else 0
            
            # Calculate total students across all groups
            total_students = len(self.reference_data)
            
            # Calculate average attendance for recorded sessions
            recorded_only = [s for s in all_sessions if s.get('status') == 'Recorded']
            avg_attendance = sum(s.get('attendance_rate', 0) for s in recorded_only) / len(recorded_only) if recorded_only else 0
            
            # Calculate statistics by group
            group_stats = {}
            for session in all_sessions:
                group_key = f"{session.get('year', '')} - {session.get('group', '')}"
                if group_key not in group_stats:
                    group_stats[group_key] = {
                        'year': session.get('year', ''),
                        'group': session.get('group', ''),
                        'total_sessions': 0,
                        'recorded_sessions': 0,
                        'missing_sessions': 0,
                        'pending_sessions': 0,
                        'total_expected_students': 0,
                        'total_attending_students': 0,
                        'expected_students_count': session.get('expected_student_count', 0)
                    }
                
                stats = group_stats[group_key]
                stats['total_sessions'] += 1
                
                status = session.get('status', '')
                if status == 'Recorded':
                    stats['recorded_sessions'] += 1
                    stats['total_expected_students'] += session.get('expected_student_count', 0)
                    stats['total_attending_students'] += session.get('unique_student_count', 0)
                elif status == 'Not Recorded':
                    stats['missing_sessions'] += 1
                elif status == 'Pending':
                    stats['pending_sessions'] += 1
            
            # Add overall class statistics table starting at column K
            stats_start_row = 2
            stats_start_col = 11  # Column K
            
            main_ws.cell(row=stats_start_row, column=stats_start_col, value="Class Statistics")
            main_ws.cell(row=stats_start_row, column=stats_start_col).font = Font(bold=True, size=14)
            
            class_stats_data = [
                ['Metric', 'Value'],
                ['Total Students', total_students],
                ['Total Sessions', total_sessions],
                ['Recorded Sessions', recorded_sessions],
                ['Missing Sessions', missing_sessions],
                ['Pending Sessions', pending_sessions],
                ['Coverage %', f"{coverage_percentage:.1f}%"],
                ['Average Attendance %', f"{avg_attendance:.1f}%"]
            ]
            
            for row_offset, row_data in enumerate(class_stats_data):
                current_row = stats_start_row + 1 + row_offset
                for col_offset, value in enumerate(row_data):
                    cell = main_ws.cell(row=current_row, column=stats_start_col + col_offset, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if row_offset == 0:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
            
            # Add group statistics table below class statistics
            group_stats_start_row = stats_start_row + len(class_stats_data) + 3  # Add spacing
            
            main_ws.cell(row=group_stats_start_row, column=stats_start_col, value="Group Statistics")
            main_ws.cell(row=group_stats_start_row, column=stats_start_col).font = Font(bold=True, size=14)
            
            # Add group stats headers
            group_stats_headers = [
                'Year', 'Group', 'Students', 'Total Sessions', 'Recorded', 
                'Missing', 'Pending', 'Coverage %', 'Avg Attendance %'
            ]
            
            current_row = group_stats_start_row + 1
            for col_offset, header in enumerate(group_stats_headers):
                cell = main_ws.cell(row=current_row, column=stats_start_col + col_offset, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add group stats data
            for group_key in sorted(group_stats.keys()):
                current_row += 1
                stats = group_stats[group_key]
                coverage = (stats['recorded_sessions'] / stats['total_sessions'] * 100) if stats['total_sessions'] > 0 else 0
                avg_attendance = (stats['total_attending_students'] / stats['total_expected_students'] * 100) if stats['total_expected_students'] > 0 else 0
                
                group_stats_row = [
                    stats['year'],
                    stats['group'],
                    stats['expected_students_count'],
                    stats['total_sessions'],
                    stats['recorded_sessions'],
                    stats['missing_sessions'],
                    stats['pending_sessions'],
                    f"{coverage:.1f}%",
                    f"{avg_attendance:.1f}%"
                ]
                
                for col_offset, value in enumerate(group_stats_row):
                    cell = main_ws.cell(row=current_row, column=stats_start_col + col_offset, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add Legend below group statistics
            legend_start_row = current_row + 3  # Add spacing after group stats
            
            main_ws.cell(row=legend_start_row, column=stats_start_col, value="Legend")
            main_ws.cell(row=legend_start_row, column=stats_start_col).font = Font(bold=True, size=14)
            
            legend_data = [
                ['Status', 'Color', 'Description'],
                ['Recorded', 'Green', 'Session recorded'],
                ['Not Recorded', 'Red', 'Session missed'],
                ['Pending', 'Yellow', 'Future session']
            ]
            
            for row_offset, row_data in enumerate(legend_data):
                current_row = legend_start_row + 1 + row_offset
                for col_offset, value in enumerate(row_data):
                    cell = main_ws.cell(row=current_row, column=stats_start_col + col_offset, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    if row_offset == 0:  # Header row
                        cell.font = header_font
                        cell.fill = header_fill
                    elif row_offset == 1:  # Green row
                        if col_offset == 1:
                            cell.fill = recorded_fill
                    elif row_offset == 2:  # Red row
                        if col_offset == 1:
                            cell.fill = not_recorded_fill
                    elif row_offset == 3:  # Yellow row
                        if col_offset == 1:
                            cell.fill = pending_fill
            
            # Auto-size columns for statistics and legend tables
            for col_num in range(stats_start_col, stats_start_col + 9):  # K to S
                max_length = 0
                column_letter = get_column_letter(col_num)
                
                for row in main_ws.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                
                adjusted_width = min(max_length + 2, 25)
                main_ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save Excel file
            wb.save(excel_filepath)
            
            # Return both file paths
            return [excel_filepath, json_filepath]
                
        except Exception as e:
            print(f"Error exporting session analysis results: {str(e)}")
            return []
        
#==========================================================Automation Coder==========================================================#

class AutomatedAttendanceProcessor:
    """
    Automated attendance processor that detects files and processes all attendance reports
    without requiring manual input through dialogs.
    """
    
    def __init__(self):
        self.base_dir = os.getcwd()
        self.reference_dir = os.path.join(self.base_dir, "reference_data")
        self.log_history_dir = os.path.join(self.base_dir, "log_history")
        self.schedules_dir = os.path.join(self.base_dir, "modules_schedules")
        self.reports_dir = os.path.join(self.base_dir, "attendance_reports")
        
        # Initialize SessionAnalyzer
        self.session_analyzer = SessionAnalyzer()
        
        # Constants
        self.ATTENDANCE_THRESHOLD = 0.75  # Hardcoded to 75%
        
        # Time window constants from UpdateAttendance
        # Note: Time windows are now calculated dynamically based on session duration
        self.STANDARD_BEFORE_MINUTES = 15
        self.STANDARD_AFTER_MINUTES = 120
        self.EXCEPTION_BEFORE_MINUTES = 15
        self.EXCEPTION_AFTER_MINUTES = 120
        self.EXCEPTION_HOURS = [12, 1, 13, 3, 15]
        
        # Minimum number of consecutive sessions to consider a group transfer confirmed
        self.TRANSFER_CONFIRMATION_THRESHOLD = 1
        
        # Define subject colors (copied from UpdateAttendance)
        self.SUBJECT_COLORS = {
            "anatomy": {"bg": "800020", "text": "FFFFFF"},
            "histology": {"bg": "FFE4E1", "text": "000000"},
            "pathology": {"bg": "663399", "text": "FFFFFF"},
            "parasitology": {"bg": "556B2F", "text": "FFFFFF"},
            "physiology": {"bg": "D4A017", "text": "FFFFFF"},
            "microbiology": {"bg": "4682B4", "text": "FFFFFF"},
            "pharmacology": {"bg": "000080", "text": "FFFFFF"},
            "biochemistry": {"bg": "1A3668", "text": "FFFFFF"},
            "clinical": {"bg": "333333", "text": "FFFFFF"},
            "other": {"bg": "000000", "text": "FFFFFF"}
        }
        
    def get_egypt_time(self):
        """Get current time in Egypt timezone (UTC+2)"""
        egypt_tz = pytz.timezone('Africa/Cairo')
        return datetime.now(egypt_tz)
        
    def normalize_date_str(self, date_val):
        """Normalize date strings to consistent DD/MM/YYYY format"""
        try:
            if isinstance(date_val, datetime):
                return date_val.strftime('%d/%m/%Y')
            elif isinstance(date_val, date):
                return date_val.strftime('%d/%m/%Y')
            elif isinstance(date_val, str):
                # Try to parse and reformat
                date_val = date_val.strip()  # Remove extra spaces
                for fmt in ['%d/%m/%Y', '%d/%#m/%Y', '%#d/%m/%Y', '%#d/%#m/%Y']:
                    try:
                        parsed = datetime.strptime(date_val, fmt)
                        return parsed.strftime('%d/%m/%Y')
                    except:
                        continue
            return str(date_val)
        except:
            return str(date_val)
        
        
    def normalize_whitespace(self, text):
        """
        Normalize whitespace in text by:
        1. Removing leading whitespace
        2. Removing trailing whitespace
        3. Collapsing multiple consecutive spaces into one
        """
        if not isinstance(text, str):
            return text
        
        # Strip leading and trailing whitespace
        text = text.strip()
        
        # Replace multiple consecutive spaces with single space
        text = ' '.join(text.split())
        
        return text
    
    def normalize_row_data(self, row):
        """
        Normalize all string values in a row by removing extra whitespace.
        Returns a new list with normalized values.
        """
        normalized_row = []
        for cell in row:
            if isinstance(cell, str):
                normalized_row.append(self.normalize_whitespace(cell))
            else:
                normalized_row.append(cell)
        return normalized_row
    
    def detect_files(self):
        """Detect all files in the directory structure and organize them by complete module name"""
        print("Detecting files in directory structure...")
        
        # Check if required directories exist
        required_dirs = [self.reference_dir, self.log_history_dir, self.schedules_dir, self.reports_dir]
        for dir_path in required_dirs:
            if not os.path.exists(dir_path):
                print(f"Warning: Directory {dir_path} does not exist")
                return None
        
        # Group files by complete module name (e.g., Y1_B2425_Introduction_to_Anatomy)
        module_groups = {}
        
        # Scan reference files
        if os.path.exists(self.reference_dir):
            for file in os.listdir(self.reference_dir):
                if file.endswith('.xlsx') and '_reference_data.xlsx' in file:
                    # Extract complete module name from reference file
                    # e.g., Y1_B2425_Introduction_to_Anatomy_reference_data.xlsx -> Y1_B2425_Introduction_to_Anatomy
                    module_name = file.replace('_reference_data.xlsx', '')
                    if module_name not in module_groups:
                        module_groups[module_name] = {}
                    module_groups[module_name]['reference'] = os.path.join(self.reference_dir, file)
        
        # Scan log history files (merge all)
        all_log_files = []
        if os.path.exists(self.log_history_dir):
            for file in os.listdir(self.log_history_dir):
                if file.endswith('.xlsx'):
                    all_log_files.append(os.path.join(self.log_history_dir, file))
        
        # Scan schedule files and match with reference files
        if os.path.exists(self.schedules_dir):
            for file in os.listdir(self.schedules_dir):
                if file.endswith('_schedule.xlsx'):
                    # Extract complete module name from schedule file
                    # e.g., Y1_B2425_Introduction_to_Anatomy_schedule.xlsx -> Y1_B2425_Introduction_to_Anatomy
                    module_name = file.replace('_schedule.xlsx', '')
                    
                    # Only add if we have a corresponding reference file
                    if module_name in module_groups:
                        module_groups[module_name]['schedule'] = os.path.join(self.schedules_dir, file)
                    else:
                        print(f"Warning: Schedule file {file} has no corresponding reference file")
        
        # Scan attendance report files and match with reference files
        if os.path.exists(self.reports_dir):
            for file in os.listdir(self.reports_dir):
                if file.endswith('_attendance.xlsx'):
                    # Extract complete module name from attendance file
                    # e.g., Y1_B2425_Introduction_to_Anatomy_attendance.xlsx -> Y1_B2425_Introduction_to_Anatomy
                    module_name = file.replace('_attendance.xlsx', '')
                    
                    # Only add if we have a corresponding reference file
                    if module_name in module_groups:
                        module_groups[module_name]['report'] = os.path.join(self.reports_dir, file)
                    else:
                        print(f"Warning: Attendance report {file} has no corresponding reference file")
        
        # Add merged log files to each module group
        for module_name in module_groups:
            module_groups[module_name]['log_files'] = all_log_files
        
        # Filter out modules that don't have both reference and schedule files
        valid_module_groups = {}
        for module_name, data in module_groups.items():
            if 'reference' in data and 'schedule' in data:
                valid_module_groups[module_name] = data
            else:
                missing_files = []
                if 'reference' not in data:
                    missing_files.append('reference')
                if 'schedule' not in data:
                    missing_files.append('schedule')
                print(f"Warning: Module {module_name} is missing required files: {', '.join(missing_files)}")
        
        print(f"Detected {len(valid_module_groups)} complete module groups:")
        for module_name, data in valid_module_groups.items():
            print(f"  {module_name}:")
            if 'reference' in data:
                print(f"    Reference: {os.path.basename(data['reference'])}")
            if 'schedule' in data:
                print(f"    Schedule: {os.path.basename(data['schedule'])}")
            if 'report' in data:
                print(f"    Report: {os.path.basename(data['report'])}")
            print(f"    Log files: {len(data['log_files'])} files")
        
        return valid_module_groups
    
    def extract_data_from_report(self, report_file):
        """Extract required data from an existing attendance report and calculate threshold"""
        try:
            wb = openpyxl.load_workbook(report_file)
            
            # Find the summary sheet
            summary_sheet = None
            for sheet_name in wb.sheetnames:
                if "Summary" in sheet_name:
                    summary_sheet = wb[sheet_name]
                    break
            
            if not summary_sheet:
                print(f"Warning: No summary sheet found in {report_file}")
                return None
            
            # Find header row
            header_row = None
            for row_idx, row in enumerate(summary_sheet.iter_rows(values_only=True)):
                if row and "Total Required" in row:
                    header_row = row_idx + 1
                    break
            
            if not header_row:
                print(f"Warning: Could not find header row in {report_file}")
                return None
            
            # Get header values to find column indices - NORMALIZE WHITESPACE
            header_values = list(summary_sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True).__next__())
            header_values = [self.normalize_whitespace(str(h)) if h else h for h in header_values]
            
            # Find column indices
            col_indices = {}
            for col_idx, cell_value in enumerate(header_values):
                normalized_value = self.normalize_whitespace(str(cell_value)) if cell_value else ""
                if normalized_value == "Total Required":
                    col_indices["total_required"] = col_idx
                elif normalized_value == "Sessions Needed":
                    col_indices["sessions_needed"] = col_idx
                elif normalized_value == "Total Attended":
                    col_indices["total_attended"] = col_idx
            
            if "total_required" not in col_indices:
                print(f"Warning: Could not find 'Total Required' column in {report_file}")
                return None
            
            # Extract total required sessions
            total_required = None
            for row in summary_sheet.iter_rows(min_row=header_row+1, values_only=True):
                if row and row[col_indices["total_required"]]:
                    total_required = int(row[col_indices["total_required"]])
                    break
            
            if total_required is None:
                print(f"Warning: Could not extract total required sessions from {report_file}")
                return None
            
            # Calculate threshold from a student row where "Sessions Needed" is not zero
            calculated_threshold = None
            best_row = None
            best_sessions_needed = float('inf')
            
            for row_idx, row in enumerate(summary_sheet.iter_rows(min_row=header_row+1, values_only=True)):
                if row and len(row) > max(col_indices.values()):
                    sessions_needed = row[col_indices.get("sessions_needed", 0)]
                    total_attended = row[col_indices.get("total_attended", 0)]
                    
                    if sessions_needed is not None and sessions_needed != 0 and total_attended is not None:
                        if sessions_needed < best_sessions_needed:
                            best_sessions_needed = sessions_needed
                            best_row = row_idx + header_row + 1
                            
                            obligatory_sessions = sessions_needed + total_attended
                            calculated_threshold = obligatory_sessions / total_required
                            
                            print(f"  Found better row (row {best_row}): Sessions Needed={sessions_needed}, Total Attended={total_attended}")
                            print(f"  Calculated threshold: {calculated_threshold:.1%} (from {obligatory_sessions}/{total_required})")
            
            if calculated_threshold is None:
                print(f"  No suitable row found for threshold calculation")
                calculated_threshold = self.ATTENDANCE_THRESHOLD
                print(f"  Using default threshold: {calculated_threshold:.1%}")
            
            return {
                'total_required': total_required,
                'threshold': calculated_threshold
            }
            
        except Exception as e:
            print(f"Error extracting data from report {report_file}: {str(e)}")
            return None
    
    def merge_log_files(self, log_files):
        """Merge all log files into a single dataset, reading ALL sheets from each workbook"""
        print(f"Merging {len(log_files)} log files...")
        
        all_log_data = []
        header_row = None
        
        for log_file in log_files:
            try:
                wb = openpyxl.load_workbook(log_file)
                
                # Iterate through ALL sheets in the workbook
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    log_data = list(ws.values)
                    
                    if not log_data:
                        continue
                    
                    # NORMALIZE WHITESPACE in header row
                    if header_row is None:
                        header_row = self.normalize_row_data(log_data[0])
                        all_log_data.append(header_row)
                    
                    # Add data rows (skip header for subsequent sheets/files)
                    # NORMALIZE WHITESPACE in all data rows
                    start_idx = 1 if len(all_log_data) > 0 else 0
                    for row in log_data[start_idx:]:
                        if row and any(cell is not None for cell in row):
                            normalized_row = self.normalize_row_data(row)
                            all_log_data.append(normalized_row)
                    
                    print(f"  Added {len(log_data) - 1} rows from {os.path.basename(log_file)} - Sheet: {sheet_name}")
                
            except Exception as e:
                print(f"Error reading log file {log_file}: {str(e)}")
        
        print(f"Total merged log data: {len(all_log_data)} rows")
        return all_log_data
    
    def create_student_map(self, student_db):
        """Create a map of student details from the reference file - WITH PROPER NORMALIZATION"""
        student_map = {}
        for row in student_db[1:]:  # Skip header
            if row[0]:
                # CRITICAL: Normalize EVERYTHING when creating the map
                student_id = self.normalize_whitespace(str(row[0]))
                name = self.normalize_whitespace(str(row[1])) if row[1] else ""
                year = self.normalize_whitespace(str(row[2])) if row[2] else ""
                # UPPERCASE normalization is CRITICAL for group matching
                group = self.normalize_whitespace(str(row[3]).upper()) if row[3] else ""
                email = f"{student_id}@med.asu.edu.eg"
                
                student_map[student_id] = {
                    "name": name,
                    "year": year,
                    "group": group,  # This MUST be normalized
                    "email": email
                }
        return student_map
    
    def extract_student_map_from_summary(self, summary_sheet):
        """Extract student information from the previous summary sheet"""
        student_map = {}
        header_row = None
        
        # Find the header row
        for row_idx, row in enumerate(summary_sheet.iter_rows(values_only=True)):
            if row and "Student ID" in row:
                header_row = row_idx + 1
                break
        
        if not header_row:
            return student_map
            
        # Find column indices - NORMALIZE WHITESPACE in headers
        col_indices = {}
        header_values = list(summary_sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True).__next__())
        for col_idx, cell_value in enumerate(header_values):
            normalized_value = self.normalize_whitespace(str(cell_value)) if cell_value else ""
            if normalized_value == "Student ID":
                col_indices["id"] = col_idx
            elif normalized_value == "Name":
                col_indices["name"] = col_idx
            elif normalized_value == "Year":
                col_indices["year"] = col_idx
            elif normalized_value == "Group":
                col_indices["group"] = col_idx
            elif normalized_value == "Email":
                col_indices["email"] = col_idx
        
        # Extract student data - NORMALIZE WHITESPACE
        for row in summary_sheet.iter_rows(min_row=header_row+1, values_only=True):
            if row and row[col_indices["id"]]:
                student_id = self.normalize_whitespace(str(row[col_indices["id"]]))
                student_map[student_id] = {
                    "name": self.normalize_whitespace(str(row[col_indices["name"]])) if row[col_indices["name"]] else "",
                    "year": row[col_indices["year"]],
                    "group": self.normalize_whitespace(str(row[col_indices["group"]])) if row[col_indices["group"]] else "",
                    "email": self.normalize_whitespace(str(row[col_indices["email"]])) if row[col_indices["email"]] else ""
                }
        
        return student_map
    
    def identify_transferred_students(self, prev_student_map, current_student_map):
        """Identify students who have changed groups"""
        transferred_students = {}
        
        for student_id, prev_data in prev_student_map.items():
            if student_id in current_student_map:
                current_data = current_student_map[student_id]
                # Check if the group has changed (both already normalized to uppercase)
                if prev_data["group"] != current_data["group"]:
                    transferred_students[student_id] = {
                        "previous_group": prev_data["group"],
                        "current_group": current_data["group"],
                        "name": current_data["name"],
                        "year": current_data["year"],
                        "email": current_data["email"]
                    }
        
        return transferred_students
    
    def extract_attendance_data(self, attendance_sheet):
        """Extract attendance data from the previous attendance sheet"""
        attendance_data = {}
        header_row = None
        
        # Find the header row
        for row_idx, row in enumerate(attendance_sheet.iter_rows(values_only=True)):
            if row and "Student ID" in row:
                header_row = row_idx + 1
                break
        
        if not header_row:
            return attendance_data
        
        # Extract attendance entries - NORMALIZE WHITESPACE
        for row in attendance_sheet.iter_rows(min_row=header_row+1, values_only=True):
            if row and len(row) >= 10:
                # Normalize the row data
                normalized_row = self.normalize_row_data(list(row))
                
                student_id = str(normalized_row[0])
                student_year = normalized_row[2]
                student_group = normalized_row[3]  # Already normalized
                
                key = f"{student_year}-{student_group}"
                
                if key not in attendance_data:
                    attendance_data[key] = []
                
                attendance_data[key].append(normalized_row)
        
        return attendance_data
    
    def calculate_completed_sessions(self, session_schedule, current_datetime):
        """Calculate completed sessions - ENSURE NORMALIZATION"""
        completed_sessions = {}
        sessions_left = {}
        
        sessions_by_group = {}
        
        for row in session_schedule:
            if len(row) >= 7:
                # CRITICAL: Normalize the row FIRST
                normalized_row = self.normalize_row_data(row)
                year, group, subject, session_num, date, start_time, duration = normalized_row[:7]
                
                # Then apply uppercase to group (already normalized by normalize_row_data)
                group = str(group).upper() if group else ""
                subject = str(subject).upper() if subject else ""
                
                # The key MUST use normalized data
                key = f"{year}-{group}"
                
                if key not in sessions_by_group:
                    sessions_by_group[key] = []
                
                # ... rest of the logic
                try:
                    session_start_datetime = self.parse_datetime(date, start_time)
                    
                    try:
                        duration_minutes = float(duration) if duration is not None else 120.0
                    except (ValueError, TypeError):
                        duration_minutes = 120.0
                    
                    from datetime import timedelta
                    session_end_datetime = session_start_datetime + timedelta(minutes=duration_minutes)
                    
                    sessions_by_group[key].append({
                        'start': session_start_datetime,
                        'end': session_end_datetime,
                        'duration': duration_minutes
                    })
                except Exception as e:
                    print(f"Error parsing date/time for session: {e}")
        
        # Count completed and remaining sessions
        for key, sessions in sessions_by_group.items():
            completed = 0
            remaining = 0
            
            for session in sessions:
                if session['end'] <= current_datetime:
                    completed += 1
                else:
                    remaining += 1
            
            completed_sessions[key] = completed
            sessions_left[key] = remaining
        
        return completed_sessions, sessions_left
    
    def calculate_required_attendance(self, session_schedule, total_required_sessions):
        """Calculate required attendance for each session by subject"""
        required_attendance = {}

        for row in session_schedule:
            if len(row) >= 4:
                # NORMALIZE FIRST
                normalized_row = self.normalize_row_data(row)
                year, group, subject, session_num = normalized_row[:4]

                # Convert to uppercase
                group = str(group).upper() if group else ""                
                subject = str(subject).upper() if subject else ""
                key = f"{year}-{group}"       

                # Initialize dictionaries if they don't exist
                if key not in required_attendance:
                    required_attendance[key] = {}
                
                if subject not in required_attendance[key]:
                    required_attendance[key][subject] = {
                        "total": 0,
                        "sessions": {}
                    }
                
                if session_num not in required_attendance[key][subject]["sessions"]:
                    required_attendance[key][subject]["sessions"][session_num] = {
                        "total": 0,
                        "locations": {}
                    }
                
                if subject not in required_attendance[key][subject]["sessions"][session_num]["locations"]:
                    required_attendance[key][subject]["sessions"][session_num]["locations"][subject] = 1
                    required_attendance[key][subject]["sessions"][session_num]["total"] += 1
                    required_attendance[key][subject]["total"] += 1
    
        return required_attendance
    
    def parse_datetime(self, date_val, time_val):
            """Parse date and time values into a timezone-aware datetime object."""
            try:
                from datetime import datetime

                # Convert date to datetime.date object
                if hasattr(date_val, 'year') and hasattr(date_val, 'month') and hasattr(date_val, 'day'):
                    if hasattr(date_val, 'hour'):
                        date_part = date_val.date()
                    else:
                        date_part = date_val
                else:
                    # NORMALIZE WHITESPACE for string dates
                    date_str = self.normalize_whitespace(str(date_val)) if isinstance(date_val, str) else str(date_val)
                    try:
                        date_part = datetime.strptime(date_str, '%d/%m/%Y').date()
                    except ValueError:
                        try:
                            date_part = datetime.strptime(date_str, '%Y-%m-%d').date()
                        except ValueError:
                            if isinstance(date_val, (int, float)) or (isinstance(date_val, str) and date_val.isdigit()):
                                date_num = int(float(date_val)) if isinstance(date_val, str) else int(date_val)
                                from datetime import timedelta
                                base_date = datetime(1899, 12, 30).date()
                                date_part = base_date + timedelta(days=date_num)
                            else:
                                raise ValueError(f"Unrecognized date format: {date_val} (type: {type(date_val)})")

                # Convert time to datetime.time object
                if hasattr(time_val, 'hour') and hasattr(time_val, 'minute'):
                    if hasattr(time_val, 'year'):
                        time_part = time_val.time()
                    else:
                        time_part = time_val
                else:
                    # NORMALIZE WHITESPACE for string times
                    time_str = self.normalize_whitespace(str(time_val)) if isinstance(time_val, str) else str(time_val)
                    try:
                        time_part = datetime.strptime(time_str, '%H:%M:%S').time()
                    except ValueError:
                        try:
                            time_part = datetime.strptime(time_str, '%H:%M').time()
                        except ValueError:
                            if isinstance(time_val, (int, float)) or (isinstance(time_val, str) and time_val.replace('.', '', 1).isdigit()):
                                time_num = float(time_val) if isinstance(time_val, str) else time_val
                                if 0 <= time_num < 1:
                                    hours = int(time_num * 24)
                                    minutes = int((time_num * 24 * 60) % 60)
                                    seconds = int((time_num * 24 * 3600) % 60)
                                    from datetime import time
                                    time_part = time(hours, minutes, seconds)
                                else:
                                    raise ValueError(f"Time value out of range (0-1): {time_val}")
                            else:
                                raise ValueError(f"Unrecognized time format: {time_val} (type: {type(time_val)})")

                # Combine date and time, then make it timezone-aware (Egypt timezone)
                naive_datetime = datetime.combine(date_part, time_part)
                egypt_tz = pytz.timezone('Africa/Cairo')
                aware_datetime = egypt_tz.localize(naive_datetime)
                
                return aware_datetime
        
            except Exception as e:
                error_msg = f"Error parsing date: {date_val} (type: {type(date_val)}) and time: {time_val} (type: {type(time_val)}). Error: {str(e)}"
                print(error_msg)
                return None

    def create_valid_logs_sheet(self, workbook, sheet_name, data):
        """Create the attendance log sheet"""
        sheet = workbook.create_sheet(sheet_name)
        header = ["Student ID", "Name", "Year", "Group", "Email",
                  "Subject", "Session", "Subject", "Date", "Time", "Validation Group"]
        sheet.append(header)

        # Apply header formatting
        for i, cell in enumerate(sheet[1]):
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet.row_dimensions[1].height = 22
        sheet.freeze_panes = 'C2'

        row_num = 2
        for key in data:
            for row_data in data[key]:
                while len(row_data) < len(header):
                    row_data.append(None)
                sheet.append(row_data[:len(header)])
                row_num += 1

        # Format date and time columns
        for col in 'I', 'J':
            for cell in sheet[col]:
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = 'DD/MM/YYYY' if col == 'I' else 'HH:MM:SS'

        # Auto-fit column widths
        for col_idx, column in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 12), 50)
                if col_idx == 2:
                    adjusted_width = max(adjusted_width, 25)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"

    def get_subject_color(self, subject_name):
        """Return background and text colors for a given subject"""
        # NORMALIZE WHITESPACE and convert to lowercase
        subject_lower = self.normalize_whitespace(subject_name).lower() if subject_name else ""

        for key in self.SUBJECT_COLORS:
            if key in subject_lower:
                return self.SUBJECT_COLORS[key]

        return self.SUBJECT_COLORS["other"]

    def calculate_min_sessions_needed(self, total_required, total_attended):
        """Calculate minimum number of sessions needed to meet attendance threshold"""
        if total_attended >= self.ATTENDANCE_THRESHOLD * total_required:
            return 0
        min_total_needed = math.ceil(self.ATTENDANCE_THRESHOLD * total_required)
        return min_total_needed - total_attended

    def create_summary_sheet(self, workbook, sheet_name, combined_attendance, required_attendance,
                            current_student_map, transferred_students, transfer_data, target_year, 
                            completed_sessions, sessions_left, total_required_sessions, batch, session_schedule):
        """
        Create a summary sheet that shows attendance statistics for each student,
        handling transferred students by validating their attendance against appropriate group schedules.
        """
        sheet = workbook.create_sheet(sheet_name)

        # Collect all subjects and their sessions
        subjects = {}
        for key, subject_data in required_attendance.items():
            for subject, data in subject_data.items():
                if subject not in subjects:
                    subjects[subject] = {"sessions": set(), "locations": set()}
                for session_num, session_data in data["sessions"].items():
                    subjects[subject]["sessions"].add(str(session_num))
                    subjects[subject]["locations"].update(session_data["locations"].keys())

        # Create header - UPDATED: Added Total Missed column
        header = ["Student ID", "Name", "Year", "Group", "Email", "Status", "Percentage",
                "Sessions Needed", "Sessions Left", "Sessions Completed", "Total Required", 
                "Total Attended", "Total Missed"]

        # Track column indices for subject coloring
        subject_column_ranges = {}
        current_col = len(header) + 1
        
        # Build a map of session dates and times from the schedule
        session_datetime_map = {}
        for row in session_schedule:
            if len(row) >= 7:
                # NORMALIZE FIRST - before any processing
                normalized_row = self.normalize_row_data(row)
                year, group, subject, session_num, date, start_time, duration = normalized_row[:7]
                
                # Convert to uppercase for consistency
                group = str(group).upper() if group else ""
                subject = str(subject).upper() if subject else ""
                key = f"{year}-{group}"
                
                if key not in session_datetime_map:
                    session_datetime_map[key] = {}
                if subject not in session_datetime_map[key]:
                    session_datetime_map[key][subject] = {}
                
                session_num_str = str(session_num)
                if session_num_str not in session_datetime_map[key][subject]:
                    # Format the date and time
                    date_str = ""
                    time_str = ""
                    try:
                        if hasattr(date, 'strftime'):
                            date_str = date.strftime('%d/%m/%Y')
                        else:
                            date_str = str(date)
                            
                        if hasattr(start_time, 'strftime'):
                            time_str = start_time.strftime('%H:%M:%S')
                        else:
                            time_str = str(start_time)
                    except:
                        date_str = str(date)
                        time_str = str(start_time)
                    
                    session_datetime_map[key][subject][session_num_str] = {
                        'date': date_str,
                        'time': time_str
                    }

        # Add subject totals and session details to header - UPDATED: Removed (Req) columns
        for subject in sorted(subjects.keys()):
            start_col = current_col
            
            header.extend(
                [f"Required {subject} (Total)", f"Attended {subject} (Total)"])
            current_col += 2
            
            for session in sorted(subjects[subject]["sessions"], 
                                key=lambda x: int(x) if x.isdigit() else x):
                for location in sorted(subjects[subject]["locations"]):
                    # Get date and time for this session from any matching group
                    date_time_info = ""
                    for group_key in session_datetime_map:
                        if subject in session_datetime_map[group_key]:
                            if session in session_datetime_map[group_key][subject]:
                                info = session_datetime_map[group_key][subject][session]
                                date_time_info = f" on {info['date']} at {info['time']}"
                                break
                        if date_time_info:
                            break
                    
                    # UPDATED: Only add (Att) column, removed (Req)
                    header.append(f"{subject} Session {session}")
                    current_col += 1
            
            subject_column_ranges[subject] = (start_col, current_col - 1)

        sheet.append(header)

        # Apply header formatting and colors
        for i, cell in enumerate(sheet[1], 1):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            for subject, (start_idx, end_idx) in subject_column_ranges.items():
                if start_idx <= i <= end_idx:
                    subject_color = self.get_subject_color(subject)
                    cell.fill = PatternFill("solid", fgColor=subject_color["bg"])
                    cell.font = Font(bold=True, color=subject_color["text"])
                    break
            if i <= len(header) - len(subject_column_ranges):
                if not cell.fill.fgColor.rgb:
                    cell.fill = PatternFill("solid", fgColor="D3D3D3")
        sheet.row_dimensions[1].height = 40
        sheet.freeze_panes = 'C2'

        # Status colors
        COLOR_PASS = "66E4A6"
        COLOR_FAIL = "FF4C4C"
        COLOR_HIGH_RISK = "FF7C7C"
        COLOR_MODERATE_RISK = "FFB97D"
        COLOR_LOW_RISK = "FFF1A6"
        COLOR_NO_RISK = "3388D5"
        COLOR_OTHER = "F0F0F0"

        # Process each student in the target year
        for student_id, student in current_student_map.items():
            if target_year in str(student['year']):
                current_key = f"{student['year']}-{student['group']}"
                group_completed = completed_sessions.get(current_key, 0)
                group_sessions_left = sessions_left.get(current_key, 0)
                total_attended = 0
                attendance_by_subject = {}

                is_transferred = student_id in transferred_students
                previous_key = None
                transfer_point = None
                if is_transferred:
                    previous_group = transferred_students[student_id]["previous_group"]
                    previous_key = f"{student['year']}-{previous_group}"
                    transfer_point = transfer_data.get(student_id, {}).get("transfer_date")

                unique_attendance_set = set()
                student_attendance = []
                if current_key in combined_attendance:
                    for entry in combined_attendance[current_key]:
                        if str(entry[0]) == str(student_id):
                            student_attendance.append(entry)
                if is_transferred and previous_key in combined_attendance:
                    for entry in combined_attendance[previous_key]:
                        if str(entry[0]) == str(student_id):
                            student_attendance.append(entry)

                # Process attendance data
                for entry in student_attendance:
                    subject = entry[5]
                    session_num = str(entry[6])
                    location = entry[7]
                    entry_date = None
                    normalized_date_str = ""
                    if len(entry) > 8 and entry[8]:
                        try:
                            if isinstance(entry[8], str):
                                entry_date = datetime.strptime(entry[8], '%d/%m/%Y')
                                normalized_date_str = entry[8]
                            elif isinstance(entry[8], datetime):
                                entry_date = entry[8]
                                normalized_date_str = entry_date.strftime('%d/%m/%Y')
                            elif hasattr(entry[8], 'year') and hasattr(entry[8], 'month') and hasattr(entry[8], 'day'):
                                if hasattr(entry[8], 'hour'):
                                    entry_date = entry[8]
                                    normalized_date_str = entry_date.strftime('%d/%m/%Y')
                                else:
                                    entry_date = datetime.combine(entry[8], datetime.min.time())
                                    normalized_date_str = entry_date.strftime('%d/%m/%Y')
                            else:
                                temp_dt = self.parse_datetime(entry[8], time(0, 0))
                                if temp_dt:
                                    entry_date = temp_dt
                                    normalized_date_str = temp_dt.strftime('%d/%m/%Y')
                            
                            # Make entry_date timezone-aware if it's naive
                            if entry_date and not entry_date.tzinfo:
                                egypt_tz = pytz.timezone('Africa/Cairo')
                                entry_date = egypt_tz.localize(entry_date)
                        except Exception as e:
                            print(f"Error parsing entry date: {entry[8]} - {str(e)}")
                            pass
                    validation_group = entry[10] if len(entry) > 10 else None
                    attendance_unique_id = f"{student_id}-{subject}-{session_num}-{normalized_date_str}"

                    should_count = False
                    if not is_transferred:
                        should_count = True
                    elif not validation_group:
                        if transfer_point and entry_date:
                            if entry_date < transfer_point:
                                should_count = True
                            else:
                                should_count = True
                        else:
                            should_count = True
                    else:
                        previous_group = transferred_students[student_id]["previous_group"]
                        current_group = student["group"]
                        if transfer_point and entry_date:
                            if entry_date < transfer_point:
                                should_count = (validation_group == previous_group)
                            else:
                                should_count = (validation_group == current_group)
                        else:
                            should_count = (validation_group == previous_group or validation_group == current_group)

                    if should_count and attendance_unique_id not in unique_attendance_set:
                        unique_attendance_set.add(attendance_unique_id)
                        if subject not in attendance_by_subject:
                            attendance_by_subject[subject] = {"total": 0, "sessions": {}}
                        if session_num not in attendance_by_subject[subject]["sessions"]:
                            attendance_by_subject[subject]["sessions"][session_num] = {"locations": {}}
                        location_key = subject.lower()
                        session_location_key = f"{session_num}-{location_key}"
                        if session_location_key not in attendance_by_subject[subject]["sessions"][session_num]["locations"]:
                            attendance_by_subject[subject]["sessions"][session_num]["locations"][location_key] = 1
                            attendance_by_subject[subject]["total"] += 1
                            total_attended += 1

                # Calculate status and color
                required_sessions = math.ceil(
                    self.ATTENDANCE_THRESHOLD * total_required_sessions)
                min_sessions_needed = max(
                    required_sessions - total_attended, 0)

                # Consider the module finished only when there are truly no sessions left
                if group_sessions_left == 0:
                    if total_attended >= required_sessions:
                        status, color = "Pass", COLOR_PASS
                    else:
                        status, color = "Fail", COLOR_FAIL
                else:
                    max_possible = total_attended + group_sessions_left
                    if max_possible < required_sessions:
                        status, color = "Fail", COLOR_FAIL
                    elif total_attended >= required_sessions:
                        status, color = "Pass", COLOR_PASS
                    else:
                        sessions_margin = group_sessions_left - min_sessions_needed
                        if sessions_margin <= 1:
                            status, color = "High Risk", COLOR_HIGH_RISK
                        elif sessions_margin <= 3:
                            status, color = "Moderate Risk", COLOR_MODERATE_RISK
                        elif sessions_margin <= 5:
                            status, color = "Low Risk", COLOR_LOW_RISK
                        else:
                            status, color = "No Risk", COLOR_NO_RISK

                percentage = total_attended / \
                    total_required_sessions if total_required_sessions > 0 else 0

                # UPDATED: Calculate total missed sessions
                total_missed = group_completed - total_attended

                # Create row data with sessions left and completed columns - UPDATED: Added total_missed
                row = [
                    student_id, student['name'], student['year'], student['group'],
                    student['email'], status, f"{percentage:.1%}", min_sessions_needed,
                    group_sessions_left, group_completed, total_required_sessions, 
                    total_attended, total_missed
                ]

                # Add subject totals and session details - UPDATED: Removed req_count
                for subject in sorted(subjects.keys()): 
                    # For required attendance, select the appropriate requirements based on transfer status
                    subject_req_total = 0  
                    subject_req_sessions = {}  

                    # For transferred students, we need to choose which requirements to use
                    if is_transferred:
                        # Always use current group requirements for the report
                        if current_key in required_attendance and subject in required_attendance[current_key]:  
                            curr_req = required_attendance[current_key][subject]  
                            subject_req_total = curr_req["total"]  

                            # Add sessions from current group
                            for session_num, session_data in curr_req["sessions"].items():
                                session_num_str = str(session_num)  # Ensure string format
                                if session_num_str not in subject_req_sessions:  
                                    subject_req_sessions[session_num_str] = {"locations": {}}  
                                for location, count in session_data["locations"].items():
                                    location_key = location.lower()  # Make lowercase for consistency
                                    if location_key not in subject_req_sessions[session_num_str]["locations"]:  
                                        subject_req_sessions[session_num_str]["locations"][location_key] = 0  
                                    subject_req_sessions[session_num_str]["locations"][location_key] = count  # Set, not add
                    else:
                        # For non-transferred students, just use current group
                        if current_key in required_attendance and subject in required_attendance[current_key]:  
                            curr_req = required_attendance[current_key][subject]  
                            subject_req_total = curr_req["total"]  

                            # Add sessions
                            for session_num, session_data in curr_req["sessions"].items():
                                session_num_str = str(session_num)  # Ensure string format
                                if session_num_str not in subject_req_sessions:  
                                    subject_req_sessions[session_num_str] = {"locations": {}}  
                                for location, count in session_data["locations"].items():
                                    location_key = location.lower()  # Make lowercase for consistency
                                    if location_key not in subject_req_sessions[session_num_str]["locations"]:  
                                        subject_req_sessions[session_num_str]["locations"][location_key] = 0  
                                    subject_req_sessions[session_num_str]["locations"][location_key] = count  

                    # Get actual attendance for this subject
                    subject_att = attendance_by_subject.get(subject, {"total": 0, "sessions": {}})  

                    # Add subject totals
                    row.extend([subject_req_total, subject_att["total"]])  

                    # Add session details - UPDATED: Only attendance count, '-' for future/ongoing sessions
                    for session in sorted(subjects[subject]["sessions"], key=lambda x: int(x) if x.isdigit() else x):  
                        for location in sorted(subjects[subject]["locations"]):  
                            location_key = location.lower()  # Make lowercase for consistency
                            att_count = subject_att.get("sessions", {}).get(session, {}).get("locations", {}).get(location_key, 0)  
                            
                            # Check if this session has ended for THIS STUDENT'S GROUP
                            session_ended = False
                            # Use the student's current group to determine the session date/time
                            student_key = current_key  # This is "{year}-{group}" for the student

                            if student_key in session_datetime_map:
                                if subject in session_datetime_map[student_key]:
                                    if session in session_datetime_map[student_key][subject]:
                                        session_info = session_datetime_map[student_key][subject][session]
                                        date_str = session_info['date']
                                        time_str = session_info['time']
                                        
                                        # Get duration from the session schedule
                                        duration_minutes = 120  # Default to 120 minutes
                                        for sched_row in session_schedule:
                                            if len(sched_row) >= 7:
                                                # CRITICAL: Normalize the schedule row before comparison
                                                normalized_sched = self.normalize_row_data(sched_row)
                                                s_year, s_group, s_subject, s_session, s_date, s_time, s_duration = normalized_sched[:7]
                                                
                                                s_group = str(s_group).upper() if s_group else ""
                                                s_subject = str(s_subject).upper() if s_subject else ""
                                                
                                                # Compare with normalized student data
                                                if (str(s_year) == str(student['year']) and 
                                                    s_group == student['group'] and 
                                                    s_subject == subject and 
                                                    str(s_session) == session):
                                                    try:
                                                        duration_minutes = float(s_duration) if s_duration is not None else 120.0
                                                    except (ValueError, TypeError):
                                                        duration_minutes = 120.0
                                                    break
                                        
                                        try:
                                            # Parse the session date and time
                                            session_date = datetime.strptime(date_str, '%d/%m/%Y')
                                            session_time = datetime.strptime(time_str, '%H:%M:%S').time()
                                            
                                            # Combine date and time to get session start datetime
                                            session_start = datetime.combine(session_date.date(), session_time)
                                            
                                            # Calculate session end time (start + duration)
                                            from datetime import timedelta
                                            session_end = session_start + timedelta(minutes=duration_minutes)
                                            
                                            # Check if session has ended (current time > session end time)
                                            current_time = self.get_egypt_time()
                                            session_ended = current_time > session_end
                                        except Exception as e:
                                            # If we can't parse the date/time, assume session has ended
                                            session_ended = True
                            else:
                                # If no session info found, assume session has ended
                                session_ended = True

                            # If session hasn't ended yet, use '-', otherwise use attendance count
                            if not session_ended:
                                row.append('-')
                            else:
                                row.append(att_count)

                sheet.append(row)

                # Apply cell formatting and colors for this row
                row_idx = sheet.max_row

                # Set row height
                sheet.row_dimensions[row_idx].height = 20

                # Define border style (only bottom border for separation between rows)
                bottom_border = Border(bottom=Side(style='thin'))

                # Apply cell formatting and colors for this row
                # Format status cell
                status_cell = sheet.cell(row=row_idx, column=6)
                status_cell.font = Font(bold=True)
                status_cell.fill = PatternFill("solid", fgColor=color)
                status_cell.alignment = Alignment(horizontal='center')
                status_cell.border = bottom_border

                # Format percentage cell
                percentage_cell = sheet.cell(row=row_idx, column=7)
                percentage_cell.number_format = '0.0%'
                percentage_cell.alignment = Alignment(horizontal='center')
                percentage_cell.border = bottom_border

                # Format sessions cells
                for col in range(8, 14):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = bottom_border

                # Apply subject-specific colors to the data cells
                for subject, (start_col, end_col) in subject_column_ranges.items():
                    subject_color = self.get_subject_color(subject)
                    for col in range(start_col, end_col + 1):
                        cell = sheet.cell(row=row_idx, column=col)
                        bg_color = self.lighten_color(subject_color["bg"])
                        cell.fill = PatternFill("solid", fgColor=bg_color)
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = bottom_border

                # Apply borders to remaining cells (Student ID, Name, Year, Group, Email)
                for col in range(1, 6):
                    cell = sheet.cell(row=row_idx, column=col)
                    cell.border = bottom_border

        # Add auto-filter to easily sort and filter data
        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(header))}1"

        # Improved column width auto-fitting
        column_widths = {}

        # First pass: Calculate max content length for each column
        for row in sheet.iter_rows():
            for cell in row:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                if cell.value:
                    # For headers (row 1), calculate based on word-wrapped text
                    if cell.row == 1:
                        # Split header by spaces and find the longest word
                        words = str(cell.value).split()
                        if words:
                            max_word_len = max(len(word) for word in words)
                            # For headers, consider both total length and longest word
                            header_width = min(max(max_word_len + 1, len(str(cell.value)) / 2), 30)
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), header_width)
                    else:
                        # For data cells, use the full text length
                        try:
                            text_len = len(str(cell.value))
                            column_widths[col_letter] = max(column_widths.get(col_letter, 0), text_len + 1)
                        except:
                            pass

        # Second pass: Apply calculated widths with constraints
        for col_letter, width in column_widths.items():
            col_idx = openpyxl.utils.column_index_from_string(col_letter)

            # Base width calculation
            adjusted_width = min(max(width, 10), 45)  # Min 10, Max 45

            # Special case for specific columns
            if col_idx == 2:  # Name column
                adjusted_width = max(adjusted_width, 25)  # Names need more space
            elif col_idx >= 14:  # Subject specific columns (adjusted due to new Total Missed column)
                adjusted_width = max(adjusted_width, 12)  # Subject columns need at least this width

            sheet.column_dimensions[col_letter].width = adjusted_width

    def create_transfer_log_sheet(self, workbook, sheet_name, transferred_students, transfer_data):
        """Create a sheet that logs all student transfers with their dates"""
        sheet = workbook.create_sheet(sheet_name)

        # Create header
        header = ["Student ID", "Name", "Year", "Group Before", "Group After", "Transfer Date"]
        sheet.append(header)

        # Apply header formatting
        for i, cell in enumerate(sheet[1]):
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet.row_dimensions[1].height = 22
        sheet.freeze_panes = 'C2'

        # Add data for each transferred student - NORMALIZE WHITESPACE
        for student_id, transfer_info in transferred_students.items():
            transfer_date = transfer_data.get(student_id, {}).get("transfer_date")
            formatted_date = ""
            if transfer_date:
                formatted_date = transfer_date.strftime('%d/%m/%Y %H:%M')

            row = [
                self.normalize_whitespace(str(student_id)),
                self.normalize_whitespace(str(transfer_info["name"])),
                transfer_info["year"],
                self.normalize_whitespace(str(transfer_info["previous_group"])),
                self.normalize_whitespace(str(transfer_info["current_group"])),
                formatted_date
            ]
            sheet.append(row)

        # Format date column
        for cell in sheet["F"][1:]:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY HH:MM'

        # Auto-fit column widths
        for col_idx, column in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 12), 50)
                if col_idx == 2:
                    adjusted_width = max(adjusted_width, 25)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"

    def lighten_color(self, hex_color, factor=0.75):
        """Lightens the given color by the factor"""
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)

        r = int(r + (255 - r) * factor)
        g = int(g + (255 - g) * factor)
        b = int(b + (255 - b) * factor)

        return f"{r:02x}{g:02x}{b:02x}".upper()

    def analyze_transfer_patterns(self, transferred_students, log_history, session_schedule, student_map):
        """Analyze attendance patterns to determine when students were transferred"""
        transfer_data = {}
        
        for student_id, transfer_info in transferred_students.items():
            # NORMALIZE WHITESPACE for group names
            previous_group = self.normalize_whitespace(str(transfer_info["previous_group"]))
            current_group = self.normalize_whitespace(str(transfer_info["current_group"]))
            student_year = transfer_info["year"]
            
            prev_group_key = f"{student_year}-{previous_group}"
            current_group_key = f"{student_year}-{current_group}"
            
            prev_group_sessions = self.create_session_map(session_schedule, prev_group_key)
            current_group_sessions = self.create_session_map(session_schedule, current_group_key)
            
            # Get attendance records for this student - NORMALIZE WHITESPACE
            student_logs = []
            for row in log_history[1:]:
                if len(row) >= 4:
                    # Normalize the student ID for comparison
                    log_student_id = self.normalize_whitespace(str(row[0]))
                    if log_student_id == self.normalize_whitespace(str(student_id)):
                        date_val = str(row[2]) if len(row) > 2 else ""
                        time_val = str(row[3]) if len(row) > 3 else ""
                        
                        if (date_val.startswith("Year") or time_val.startswith("C") or 
                            not date_val or not time_val):
                            continue
                        
                        student_logs.append(row)
            
            def safe_sort_key(x):
                try:
                    if len(x) >= 4:
                        date_val = x[2]
                        time_val = x[3]
                        if (isinstance(date_val, str) and date_val.startswith("Year")):
                            return datetime.min
                        if (isinstance(time_val, str) and time_val.startswith("C")):
                            return datetime.min
                        parsed = self.parse_datetime(date_val, time_val)
                        return parsed if parsed else datetime.min
                    return datetime.min
                except Exception:
                    return datetime.min
            
            student_logs.sort(key=safe_sort_key)
            
            attendance_pattern = []
            for log in student_logs:
                if len(log) < 4:
                    continue
                    
                log_datetime = self.parse_datetime(log[2], log[3])
                if not log_datetime:
                    continue
                    
                # NORMALIZE WHITESPACE for location
                location = self.normalize_whitespace(str(log[1]))
                
                prev_match = self.match_log_to_session(log, log_datetime, location, prev_group_sessions)
                current_match = self.match_log_to_session(log, log_datetime, location, current_group_sessions)
                
                if prev_match and current_match:
                    attendance_pattern.append(("both", log_datetime, location))
                elif prev_match:
                    attendance_pattern.append(("previous", log_datetime, location))
                elif current_match:
                    attendance_pattern.append(("current", log_datetime, location))
            
            transfer_point = self.detect_transfer_point(attendance_pattern)
            
            transfer_data[student_id] = {
                "previous_group": previous_group,
                "current_group": current_group,
                "transfer_date": transfer_point,
                "attendance_pattern": attendance_pattern
            }
        
        return transfer_data
    
    def create_session_map(self, sessions, group_key):
        """Create a map of sessions for a specific group"""
        session_map = {}
        
        for session in sessions:
            if len(session) >= 7:
                # NORMALIZE WHITESPACE in session data
                normalized_row = self.normalize_row_data(session)
                year, group, subject, session_num, date, start_time, duration = normalized_row[:7]

                group = str(group).upper() if group else ""
                subject = str(subject).upper() if subject else ""
                key = f"{year}-{group}"

                if key == group_key:
                    session_datetime = self.parse_datetime(date, start_time)
                    if not session_datetime:
                        continue
                    
                    try:
                        session_duration = float(duration) if duration is not None else 120.0
                    except (ValueError, TypeError):
                        session_duration = 120.0
                        
                    session_key = f"{subject}-{date}-{start_time}"
                    
                    session_map[session_key] = {
                        "subject": subject,
                        "session_num": session_num,
                        "start_time": session_datetime,
                        "date": date,
                        "duration": session_duration
                    }
        
        return session_map
    
    def match_log_to_session(self, log, log_datetime, location, session_map):
        """Check if a log matches any session in the given session map"""
        # NORMALIZE WHITESPACE for location comparison
        normalized_location = self.normalize_whitespace(str(location).lower())
        
        for session_key, session_info in session_map.items():
            session_subject = self.normalize_whitespace(str(session_info["subject"]).lower())
            session_start = session_info["start_time"]
            
            if normalized_location == session_subject:
                session_duration = session_info.get("duration", 120.0)
                before_window = timedelta(minutes=15)
                after_window = timedelta(minutes=int(session_duration))
                
                if session_start - before_window <= log_datetime <= session_start + after_window:
                    return True
        
        return False
    
    def detect_transfer_point(self, attendance_pattern):
        """Detect the point at which a student consistently switched to the new group"""
        if not attendance_pattern:
            return None

        consecutive_current = 0
        consecutive_start_idx = -1

        for idx, (group, log_datetime, _) in enumerate(attendance_pattern):
            if group == "current":
                if consecutive_current == 0:
                    consecutive_start_idx = idx
                consecutive_current += 1
        
                if consecutive_current >= self.TRANSFER_CONFIRMATION_THRESHOLD:
                    first_date_in_sequence = attendance_pattern[consecutive_start_idx][1]
                    if first_date_in_sequence is not None:
                        transfer_date = first_date_in_sequence - timedelta(days=1)
                        return transfer_date
                    else:
                        for i in range(consecutive_start_idx + 1, min(len(attendance_pattern), consecutive_start_idx + self.TRANSFER_CONFIRMATION_THRESHOLD)):
                            if attendance_pattern[i][1] is not None:
                                transfer_date = attendance_pattern[i][1] - timedelta(days=1)
                                return transfer_date
                        return None
            else:
                consecutive_current = 0
                consecutive_start_idx = -1

        return None

    def validate_attendance_with_transfers(self, log_history, session_schedule, student_map, 
                                    transferred_students, transfer_data, target_year):
        """Validate attendance considering student transfers - FIXED NORMALIZATION"""
        valid_attendance = {}
        session_map = {}
        unique_logs = set()

        # Build session maps for all groups - WITH NORMALIZATION
        for row in session_schedule:
            if len(row) >= 7:
                # CRITICAL: Normalize FIRST
                normalized_row = self.normalize_row_data(row)
                year, group, subject, session_num, date, start_time, duration = normalized_row[:7]
                
                # Apply uppercase (data already normalized)
                group = str(group).upper() if group else ""
                subject = str(subject).upper() if subject else ""
                
                # Key MUST be built from normalized data
                key = f"{year}-{group}"
                
                session_datetime = self.parse_datetime(date, start_time)
                if not session_datetime:
                    continue

                session_num = str(session_num)
                
                try:
                    session_duration = float(duration) if duration is not None else 120.0
                except (ValueError, TypeError):
                    session_duration = 120.0

                session_key = f"{subject}-{date}-{start_time}"

                if key not in session_map:
                    session_map[key] = {}

                session_map[key][session_key] = {
                    "subject": subject,
                    "session_num": session_num,
                    "start_time": session_datetime,
                    "date": date,
                    "duration": session_duration
                }

        # Process log history - WITH NORMALIZATION
        for row in log_history[1:]:
            if len(row) >= 4:
                # CRITICAL: Normalize FIRST
                normalized_row = self.normalize_row_data(row)
                student_id = str(normalized_row[0])
                location = str(normalized_row[1]).upper() if normalized_row[1] else ""
                date = normalized_row[2]
                time = normalized_row[3]

                if student_id in student_map:
                    student = student_map[student_id]
                    # Student data is already normalized when map was created
                    
                    log_datetime = self.parse_datetime(date, time)
                    if not log_datetime:
                        continue
                    
                    normalized_date = log_datetime.strftime('%d/%m/%Y')
                    
                    # Determine which group to validate against
                    if student_id in transferred_students:
                        transfer_info = transfer_data.get(student_id, {})
                        transfer_date = transfer_info.get("transfer_date")
            
                        if transfer_date and log_datetime < transfer_date:
                            group_to_use = transferred_students[student_id]["previous_group"]
                        elif transfer_date and log_datetime == transfer_date:
                            group_to_use = "both"
                        else:
                            group_to_use = student["group"]  # Already normalized
                    else:
                        group_to_use = student["group"]  # Already normalized
        
                    # Build keys - ALL data is now normalized
                    actual_key = f"{student['year']}-{student['group']}"
                    previous_key = None
                
                    if student_id in transferred_students:
                        previous_group = transferred_students[student_id]["previous_group"]
                        previous_key = f"{student['year']}-{previous_group}"
        
                    validation_keys = []
                
                    if group_to_use == "both":
                        if previous_key:
                            validation_keys.append((previous_key, transferred_students[student_id]["previous_group"]))
                        validation_keys.append((actual_key, student["group"]))
                    elif student_id in transferred_students and group_to_use == transferred_students[student_id]["previous_group"]:
                        validation_keys.append((previous_key, group_to_use))
                    else:
                        validation_keys.append((actual_key, group_to_use))
        
                    # Validate against session map - ALL keys are normalized
                    for validation_key, validation_group_name in validation_keys:
                        if validation_key in session_map:
                            for session_key, session_info in session_map[validation_key].items():
                                session_subject = session_info["subject"]
                                session_start = session_info["start_time"]
                        
                                if location.lower() == session_subject.lower():
                                    session_duration = session_info.get("duration", 120.0)
                                    before_window = timedelta(minutes=15)
                                    after_window = timedelta(minutes=int(session_duration))
                            
                                    if session_start - before_window <= log_datetime <= session_start + after_window:
                                        unique_log_key = f"{student_id}-{session_info['subject']}-{session_info['session_num']}-{location}-{normalized_date}"
                                
                                        if unique_log_key not in unique_logs:
                                            unique_logs.add(unique_log_key)
                                    
                                            if actual_key not in valid_attendance:
                                                valid_attendance[actual_key] = []
                                        
                                            valid_attendance[actual_key].append([
                                                student_id, student['name'], student['year'],
                                                student['group'], student['email'], session_info['subject'],
                                                session_info['session_num'], location, normalized_date, time,
                                                validation_group_name
                                            ])
                                            break

        return valid_attendance

    def combine_attendance_data(self, prev_attendance, new_attendance, transferred_students, transfer_data):
        """Combine previous and new attendance data, considering student transfers"""
        combined_attendance = {}
        unique_attendance_entries = set()

        # Add previous attendance data - NORMALIZE WHITESPACE
        for key, entries in prev_attendance.items():
            combined_attendance[key] = []
            for entry in entries:
                if len(entry) >= 10:
                    # Normalize the entry data
                    normalized_entry = self.normalize_row_data(entry)
                    student_id = str(normalized_entry[0])
                    subject = normalized_entry[5]
                    session_num = normalized_entry[6]
                    location = normalized_entry[7]
                    date = normalized_entry[8]
                
                    normalized_date = self.normalize_date_str(date)
                    unique_id = f"{student_id}-{subject}-{session_num}-{location}-{normalized_date}"
                
                    if unique_id not in unique_attendance_entries:
                        unique_attendance_entries.add(unique_id)
                    
                        if student_id and student_id in transferred_students:
                            new_entry = list(normalized_entry)
                            while len(new_entry) < 11:
                                new_entry.append(None)
                            if new_entry[10] is None:
                                new_entry[10] = transferred_students[student_id]["previous_group"]
                            combined_attendance[key].append(new_entry)
                        else:
                            combined_attendance[key].append(list(normalized_entry))

        # Add new attendance data - NORMALIZE WHITESPACE
        for key, entries in new_attendance.items():
            if key not in combined_attendance:
                combined_attendance[key] = []

            for entry in entries:
                if len(entry) >= 10:
                    # Normalize the entry data
                    normalized_entry = self.normalize_row_data(entry)
                    student_id = str(normalized_entry[0])
                    subject = normalized_entry[5]
                    session_num = normalized_entry[6]
                    location = normalized_entry[7]
                    date = normalized_entry[8]
                
                    normalized_date = self.normalize_date_str(date)
                    unique_id = f"{student_id}-{subject}-{session_num}-{location}-{normalized_date}"
                
                    if unique_id not in unique_attendance_entries:
                        unique_attendance_entries.add(unique_id)
                        combined_attendance[key].append(normalized_entry)

        return combined_attendance

    def extract_existing_transfers(self, transfer_sheet):
        """Extract existing transfer records from the transfer sheet"""
        existing_transfers = {}
        header_row = None

        # Find the header row - NORMALIZE WHITESPACE
        for row_idx, row in enumerate(transfer_sheet.iter_rows(values_only=True)):
            if row:
                normalized_row = self.normalize_row_data(list(row))
                if "Student ID" in normalized_row:
                    header_row = row_idx + 1
                    break

        if not header_row:
            return existing_transfers

        # Find column indices - NORMALIZE WHITESPACE
        col_indices = {}
        header_values = list(transfer_sheet.iter_rows(min_row=header_row, max_row=header_row, values_only=True).__next__())
        normalized_headers = self.normalize_row_data(header_values)
        
        for col_idx, cell_value in enumerate(normalized_headers):
            if cell_value == "Student ID":
                col_indices["id"] = col_idx
            elif cell_value == "Name":
                col_indices["name"] = col_idx
            elif cell_value == "Year":
                col_indices["year"] = col_idx
            elif cell_value == "Group Before":
                col_indices["group_before"] = col_idx
            elif cell_value == "Group After":
                col_indices["group_after"] = col_idx
            elif cell_value == "Transfer Date":
                col_indices["transfer_date"] = col_idx

        # Extract existing transfer data - NORMALIZE WHITESPACE
        for row in transfer_sheet.iter_rows(min_row=header_row+1, values_only=True):
            if row and row[col_indices["id"]]:
                normalized_row = self.normalize_row_data(list(row))
                student_id = str(normalized_row[col_indices["id"]])
                transfer_date = None
                
                if len(normalized_row) > col_indices["transfer_date"] and normalized_row[col_indices["transfer_date"]]:
                    try:
                        date_val = normalized_row[col_indices["transfer_date"]]
                        if isinstance(date_val, str):
                            transfer_date = datetime.strptime(date_val, '%d/%m/%Y %H:%M')
                        elif isinstance(date_val, datetime):
                            transfer_date = date_val
                        elif hasattr(date_val, 'year'):
                            transfer_date = date_val
                    except Exception as e:
                        print(f"Error parsing existing transfer date for {student_id}: {e}")

                existing_transfers[student_id] = {
                    "previous_group": normalized_row[col_indices["group_before"]],
                    "current_group": normalized_row[col_indices["group_after"]],
                    "name": normalized_row[col_indices["name"]],
                    "year": normalized_row[col_indices["year"]],
                    "email": f"{student_id}@med.asu.edu.eg",
                    "transfer_date": transfer_date,
                    "is_from_previous_report": True
                }

        return existing_transfers

    def combine_transfer_data(self, existing_transfers, new_transfers, current_student_map):
        """Combine existing and new transfer data, avoiding duplicates"""
        all_transfers = {}

        # Add all existing transfers - NORMALIZE WHITESPACE
        for student_id, transfer_info in existing_transfers.items():
            if student_id in current_student_map:
                current_info = current_student_map[student_id]
                all_transfers[student_id] = {
                    "previous_group": self.normalize_whitespace(str(transfer_info["previous_group"])),
                    "current_group": self.normalize_whitespace(str(current_info["group"])),
                    "name": self.normalize_whitespace(str(current_info["name"])),
                    "year": current_info["year"],
                    "email": self.normalize_whitespace(str(current_info["email"])),
                    "transfer_date": transfer_info["transfer_date"],
                    "is_from_previous_report": True
                }
            else:
                all_transfers[student_id] = {
                    "previous_group": self.normalize_whitespace(str(transfer_info["previous_group"])),
                    "current_group": self.normalize_whitespace(str(transfer_info["current_group"])),
                    "name": self.normalize_whitespace(str(transfer_info["name"])),
                    "year": transfer_info["year"],
                    "email": self.normalize_whitespace(str(transfer_info["email"])),
                    "transfer_date": transfer_info["transfer_date"],
                    "is_from_previous_report": True
                }

        # Add new transfers not already tracked - NORMALIZE WHITESPACE
        for student_id, transfer_info in new_transfers.items():
            if student_id not in all_transfers:
                all_transfers[student_id] = {
                    "previous_group": self.normalize_whitespace(str(transfer_info["previous_group"])),
                    "current_group": self.normalize_whitespace(str(transfer_info["current_group"])),
                    "name": self.normalize_whitespace(str(transfer_info["name"])),
                    "year": transfer_info["year"],
                    "email": self.normalize_whitespace(str(transfer_info["email"])),
                    "transfer_date": None,
                    "is_from_previous_report": False
                }
            else:
                existing_info = all_transfers[student_id]
                normalized_current = self.normalize_whitespace(str(transfer_info["current_group"]))
                if existing_info["current_group"] != normalized_current:
                    all_transfers[student_id]["current_group"] = normalized_current

        return all_transfers

    def create_comprehensive_transfer_log_sheet(self, workbook, sheet_name, existing_transfers,
                                               new_transfers, transfer_data):
        """Create transfer sheet with both existing and new transfers"""
        sheet = workbook.create_sheet(sheet_name)
        header = ["Student ID", "Name", "Year", "Group Before", "Group After", "Transfer Date", "Status"]
        sheet.append(header)

        for i, cell in enumerate(sheet[1]):
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D3D3D3")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        sheet.row_dimensions[1].height = 22
        sheet.freeze_panes = 'C2'

        processed_students = set()

        # Add existing transfers - NORMALIZE WHITESPACE
        for student_id, transfer_info in existing_transfers.items():
            if student_id not in processed_students:
                transfer_date = transfer_info.get("transfer_date")
                formatted_date = ""
                if transfer_date:
                    formatted_date = transfer_date.strftime('%d/%m/%Y %H:%M')
                row = [
                    self.normalize_whitespace(str(student_id)),
                    self.normalize_whitespace(str(transfer_info["name"])),
                    transfer_info["year"],
                    self.normalize_whitespace(str(transfer_info["previous_group"])),
                    self.normalize_whitespace(str(transfer_info["current_group"])),
                    formatted_date,
                    "Previous Transfer"
                ]
                sheet.append(row)
                processed_students.add(student_id)

        # Add new transfers - NORMALIZE WHITESPACE
        for student_id, transfer_info in new_transfers.items():
            if student_id not in processed_students:
                transfer_date = transfer_data.get(student_id, {}).get("transfer_date")
                formatted_date = ""
                if transfer_date:
                    formatted_date = transfer_date.strftime('%d/%m/%Y %H:%M')
                row = [
                    self.normalize_whitespace(str(student_id)),
                    self.normalize_whitespace(str(transfer_info["name"])),
                    transfer_info["year"],
                    self.normalize_whitespace(str(transfer_info["previous_group"])),
                    self.normalize_whitespace(str(transfer_info["current_group"])),
                    formatted_date,
                    "New Transfer"
                ]
                sheet.append(row)
                processed_students.add(student_id)

        for cell in sheet["F"][1:]:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY HH:MM'

        for col_idx, column in enumerate(sheet.columns, 1):
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            if max_length > 0:
                adjusted_width = min(max(max_length + 2, 12), 50)
                if col_idx == 2:
                    adjusted_width = max(adjusted_width, 25)
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = adjusted_width

        sheet.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(sheet.max_column)}{sheet.max_row}"

    def process_all_reports(self):
        """Main method to process all attendance reports automatically"""
        print("Starting automated attendance processing...")

        module_groups = self.detect_files()
        if not module_groups:
            print("No files detected or required directories missing. Exiting.")
            return

        all_log_files = []
        for module_name, data in module_groups.items():
            all_log_files.extend(data.get('log_files', []))

        all_log_files = list(dict.fromkeys(all_log_files))
        print(f"Found {len(all_log_files)} unique log files to process")
        
        # MERGE LOG FILES ONCE AT THE BEGINNING
        print("Merging all log files (this will be done only once)...")
        merged_log_data = self.merge_log_files(all_log_files)

        if not merged_log_data:
            print("No log data found. Exiting.")
            return

        print(f"Successfully merged {len(merged_log_data)} rows of log data\n")

        for module_name, data in module_groups.items():
            print(f"\nProcessing {module_name}...")

            year, batch, module_parsed = self.parse_module_info(module_name)

            reference_file = data.get('reference')
            if not reference_file:
                print(f"No reference file found for {module_name}. Skipping.")
                continue

            try:
                ref_wb = openpyxl.load_workbook(reference_file)
                ref_ws = ref_wb.active
                student_db = list(ref_ws.values)
                current_student_map = self.create_student_map(student_db)
            except Exception as e:
                print(f"Error loading reference file {reference_file}: {str(e)}")
                continue

            schedule_file = data.get('schedule')
            if not schedule_file:
                print(f"No schedule file found for {module_name}. Skipping.")
                continue

            try:
                sched_wb = openpyxl.load_workbook(schedule_file)
                sched_ws = sched_wb.active
                session_schedule = list(sched_ws.values)
            except Exception as e:
                print(f"Error loading schedule file {schedule_file}: {str(e)}")
                continue

            session_analysis_files = self.run_session_analysis(
                module_name, reference_file, schedule_file, merged_log_data, year, batch
            )

            report_file = data.get('report')
            if report_file:
                report_data = self.extract_data_from_report(report_file)
                if report_data:
                    total_required = report_data['total_required']
                    calculated_threshold = report_data['threshold']
                    self.ATTENDANCE_THRESHOLD = calculated_threshold
                    print(f"  Using calculated threshold: {calculated_threshold:.1%}")
                else:
                    print(f"Could not extract data from {report_file}. Using default threshold.")
                    total_required = 100
                    self.ATTENDANCE_THRESHOLD = 0.75
            else:
                print(f"No existing report found for {module_name}. Using default values.")
                total_required = 100
                self.ATTENDANCE_THRESHOLD = 0.75

            prev_student_map = {}
            prev_attendance_data = {}
            existing_transfers = {}

            if report_file:
                try:
                    prev_report_wb = openpyxl.load_workbook(report_file)
                    prev_summary_sheet = None
                    prev_attendance_sheet = None
                    prev_transfer_sheet = None

                    for sheet_name in prev_report_wb.sheetnames:
                        if "Summary" in sheet_name:
                            prev_summary_sheet = prev_report_wb[sheet_name]
                        elif "Attendance" in sheet_name:
                            prev_attendance_sheet = prev_report_wb[sheet_name]
                        elif "Transfer" in sheet_name:
                            prev_transfer_sheet = prev_report_wb[sheet_name]

                    if prev_summary_sheet:
                        prev_student_map = self.extract_student_map_from_summary(prev_summary_sheet)
                    if prev_attendance_sheet:
                        prev_attendance_data = self.extract_attendance_data(prev_attendance_sheet)
                    if prev_transfer_sheet:
                        existing_transfers = self.extract_existing_transfers(prev_transfer_sheet)
                        
                    print(f"  Loaded previous data: {len(prev_student_map)} students, {len(existing_transfers)} transfers")
                except Exception as e:
                    print(f"Error loading previous report data: {str(e)}")

            # Identify transferred students
            new_transferred_students = self.identify_transferred_students(prev_student_map, current_student_map)
            all_transferred_students = self.combine_transfer_data(existing_transfers, new_transferred_students, current_student_map)
            
            if new_transferred_students:
                print(f"  Found {len(new_transferred_students)} new transferred students")
            if all_transferred_students:
                print(f"  Total transferred students: {len(all_transferred_students)}")

            # Calculate session completion status
            current_datetime = self.get_egypt_time()
            completed_sessions, sessions_left = self.calculate_completed_sessions(
                session_schedule[1:], current_datetime)
            required_attendance = self.calculate_required_attendance(session_schedule[1:], total_required)

            # Analyze transfer patterns using the merged log data
            transfer_data = self.analyze_transfer_patterns(all_transferred_students, merged_log_data,
                                                        session_schedule[1:], current_student_map)

            # Validate attendance using the merged log data
            new_valid_attendance = self.validate_attendance_with_transfers(
                merged_log_data, session_schedule[1:], current_student_map,
                all_transferred_students, transfer_data, f"Year {year}")

            # Combine with previous attendance data
            combined_attendance = self.combine_attendance_data(
                prev_attendance_data, new_valid_attendance, all_transferred_students, transfer_data)

            print(f"  Processed attendance for {sum(len(entries) for entries in combined_attendance.values())} records")

            # Create new workbook and sheets
            output_wb = openpyxl.Workbook()
            output_wb.remove(output_wb.active)

            summary_sheet_name = f"Summary"
            attendance_sheet_name = f"Attendance"
            transfer_sheet_name = f"Transfers"

            # Create sheets
            self.create_summary_sheet(output_wb, summary_sheet_name, combined_attendance, 
                                    required_attendance, current_student_map, all_transferred_students,
                                    transfer_data, f"Year {year}", completed_sessions, sessions_left, 
                                    total_required, batch, session_schedule[1:])  # Add session_schedule parameter

            self.create_valid_logs_sheet(output_wb, attendance_sheet_name, combined_attendance)

            if all_transferred_students:
                self.create_comprehensive_transfer_log_sheet(output_wb, transfer_sheet_name,
                                                            existing_transfers, new_transferred_students,
                                                            transfer_data)

            # Save the workbook
            output_filename = f"{module_name}_attendance.xlsx"
            output_path = os.path.join(self.reports_dir, output_filename)
            output_wb.save(output_path)

            # Set Excel metadata
            try:
                wb = openpyxl.load_workbook(output_path)
                props = wb.properties
                props.title = "attendance_reports"
                props.subject = f"{module_name}_attendance"
                props.authors = f"{module_name}"
                wb.save(output_path)
            except Exception as meta_exc:
                print(f"Warning: Could not set Excel metadata: {meta_exc}")

            # Create JSON version
            try:
                df = pd.read_excel(output_path)
                json_path = os.path.splitext(output_path)[0] + '.json'
                metadata = {
                    "title": "attendance_reports",
                    "subject": f"{module_name}_attendance",
                    "authors": f"{module_name}",
                    "processing_timestamp": datetime.now().isoformat(),
                    "total_log_files_processed": len(all_log_files),
                    "session_analysis_files": [os.path.basename(f) for f in session_analysis_files] if session_analysis_files else []
                }
                json_obj = {
                    "metadata": metadata,
                    "attendance_data": df.to_dict(orient='records')
                }
                import json as _json
                with open(json_path, 'w', encoding='utf-8') as f:
                    _json.dump(json_obj, f, ensure_ascii=False, indent=2)
                print(f"  Successfully created {output_filename} and {os.path.basename(json_path)}")
                
                # Report session analysis results
                if session_analysis_files:
                    print(f"  Session analysis files: {[os.path.basename(f) for f in session_analysis_files]}")
                    
            except Exception as e:
                print(f"  Error converting to JSON: {str(e)}")

        print(f"\nAutomated attendance processing completed!")
        print(f"Processed {len(module_groups)} modules using {len(all_log_files)} log files")
        print(f"All reports saved to: {self.reports_dir}")
        
    def parse_module_info(self, module_name):
        """Parse module name to extract year, batch, and module components"""
        try:
            parts = module_name.split('_')
            if len(parts) >= 2:
                year_part = parts[0].replace('Y', '')  # Remove Y prefix
                batch_part = parts[1].replace('B', '')  # Remove B prefix
                
                # Join remaining parts as module name (everything after Y and B parts)
                if len(parts) > 2:
                    module_part = '_'.join(parts[2:])
                else:
                    module_part = "Unknown_Module"
                    
                return year_part, batch_part, module_part
            else:
                return "Unknown", "Unknown", "Unknown_Module"
        except Exception as e:
            print(f"Error parsing module name {module_name}: {str(e)}")
            return "Unknown", "Unknown", "Unknown_Module"
        
    def run_session_analysis(self, module_name, reference_file, schedule_file, merged_log_data, year, batch):
        """Run session analysis for a specific module using pre-merged log data"""
        try:
            print(f"  Running session analysis for {module_name}...")
            
            if not merged_log_data:
                print(f"  No merged log data available for session analysis")
                return []
            
            # Parse module info for consistent naming
            year_parsed, batch_parsed, module_parsed = self.parse_module_info(module_name)
            
            # Get the actual sheet names from each file
            def get_first_sheet_name(file_path):
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    return wb.sheetnames[0] if wb.sheetnames else 'Sheet1'
                except Exception as e:
                    print(f"    Warning: Could not read sheet names from {file_path}: {e}")
                    return 'Sheet1'  # Fallback
            
            # Get actual sheet names
            ref_sheet_name = get_first_sheet_name(reference_file)
            schedule_sheet_name = get_first_sheet_name(schedule_file)
            
            print(f"    Using sheets - Reference: '{ref_sheet_name}', Schedule: '{schedule_sheet_name}'")
            print(f"    Using pre-merged log data ({len(merged_log_data)} rows)")
            
            # Run the analysis using the SessionAnalyzer with merged log data
            success, message, generated_files = self.session_analyzer.analyze_sessions_with_merged_logs(
                ref_file_path=reference_file,
                ref_sheet=ref_sheet_name,
                merged_log_data=merged_log_data,  # Pass the merged data directly
                schedule_file_path=schedule_file,
                schedule_sheet=schedule_sheet_name,
                year=year_parsed,
                batch=batch_parsed,
                module=module_parsed
            )
            
            if success:
                print(f"  Session analysis completed: {message}")
                if generated_files:
                    for file in generated_files:
                        print(f"    Generated: {os.path.basename(file)}")
                return generated_files
            else:
                print(f"  Session analysis failed: {message}")
                return []
                
        except Exception as e:
            print(f"  Error during session analysis: {str(e)}")
            return []


if __name__ == "__main__":
    # Check if we should run the automated processor
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--auto":
        # Run automated processor
        processor = AutomatedAttendanceProcessor()
        processor.process_all_reports()
    else:
        # For now, just run the automated processor
        print("Running automated attendance processor...")
        processor = AutomatedAttendanceProcessor()
        processor.process_all_reports()
        
    